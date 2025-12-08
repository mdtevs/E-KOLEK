from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.core.serializers import serialize
from django.utils import timezone
from django.db import transaction
from django.db.models import Max
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
import json
import pandas as pd
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from cenro.admin_auth import admin_required
from .models import LearningVideo, VideoWatchHistory, QuizQuestion, QuizResult, QuizAnswer
from .serializers import (
    LearningVideoSerializer, QuizQuestionSerializer, QuizSubmissionSerializer, 
    QuizResultSerializer, QuizAnswerInputSerializer
)
from accounts.models import PointsTransaction

logger = logging.getLogger(__name__)

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def get_learning_videos(request):
    """Get all active learning videos with quiz information"""
    try:
        # Get all active videos
        videos = LearningVideo.objects.filter(is_active=True)
        
        # Use serializer for consistent data format
        serializer = LearningVideoSerializer(videos, many=True, context={'request': request})
        
        # Debug info
        total_videos = LearningVideo.objects.count()
        active_videos = videos.count()
        
        return Response({
            'success': True,
            'videos': serializer.data,
            'debug_info': {
                'total_videos_in_db': total_videos,
                'active_videos_count': active_videos,
                'user_authenticated': request.user.is_authenticated,
                'user_id': request.user.id if request.user.is_authenticated else None
            }
        })
    
    except Exception as e:
        return Response({
            'success': False,
            'message': f'Error fetching videos: {str(e)}',
            'error_type': type(e).__name__
        }, status=500)

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def mark_video_as_watched(request):
    """Mark a video as watched and award points"""
    try:
        data = request.data if hasattr(request, 'data') else json.loads(request.body)
        video_id = data.get('video_id')
        
        if not video_id:
            return JsonResponse({
                'success': False,
                'error': 'Video ID is required'
            }, status=400)
        
        # Get the video
        video = LearningVideo.objects.get(id=video_id, is_active=True)
        points_earned = video.points_reward
        
        # Debug info before processing
        debug_info = {
            'user_id': request.user.id,
            'user_authenticated': request.user.is_authenticated,
            'video_id': video_id,
            'video_title': video.title,
            'points_to_award': points_earned,
            'user_points_before': request.user.total_points,
            'family_exists': bool(request.user.family),
        }
        
        # Check if user has already watched this video
        watch_history, created = VideoWatchHistory.objects.get_or_create(
            user=request.user,
            video=video,
            defaults={'points_earned': points_earned}
        )
        
        debug_info['watch_history_created'] = created
        debug_info['already_watched'] = not created
        
        if created:
            # Store original points for comparison
            original_user_points = request.user.total_points
            original_family_points = request.user.family.total_family_points if request.user.family else 0
            
            # Award points to user
            request.user.total_points += points_earned
            request.user.save()
            
            # Award points to family
            if request.user.family:
                request.user.family.total_family_points += points_earned
                request.user.family.save()
                debug_info['family_points_before'] = original_family_points
                debug_info['family_points_after'] = request.user.family.total_family_points
            
            # Create points transaction record
            transaction = PointsTransaction.objects.create(
                user=request.user,
                transaction_type='earned',
                points_amount=points_earned,
                description=f"Watched learning video: {video.title}",
                transaction_date=watch_history.watched_at
            )
            
            # Create notification for learning video completion
            from accounts.models import Notification
            Notification.objects.create(
                user=request.user,
                type='learning',
                message=f'You earned {points_earned} points for watching "{video.title}"!',
                points=points_earned,
                video_title=video.title
            )
            
            debug_info['user_points_after'] = request.user.total_points
            debug_info['transaction_created'] = transaction.id
            debug_info['points_actually_added'] = request.user.total_points - original_user_points
            
            return JsonResponse({
                'success': True,
                'new_total_points': request.user.total_points,
                'points_earned': points_earned,
                'message': f'Congratulations! You earned {points_earned} points!',
                'debug_info': debug_info
            })
        else:
            debug_info['watch_date'] = watch_history.watched_at.isoformat() if watch_history.watched_at else None
            debug_info['points_previously_earned'] = watch_history.points_earned
            
            return JsonResponse({
                'success': True,
                'new_total_points': request.user.total_points,
                'points_earned': 0,
                'message': 'You have already watched this video.',
                'debug_info': debug_info
            })
    
    except LearningVideo.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Video not found'
        }, status=404)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error marking video as watched: {str(e)}'
        }, status=500)

@require_http_methods(["GET"])
def debug_learning_data(request):
    """Debug endpoint to check learning videos and user data"""
    try:
        # Get all learning videos
        all_videos = LearningVideo.objects.all()
        active_videos = LearningVideo.objects.filter(is_active=True)
        
        # Get user info if authenticated
        user_info = None
        watch_history = []
        if request.user.is_authenticated:
            user_info = {
                'id': request.user.id,
                'username': getattr(request.user, 'username', 'N/A'),
                'total_points': request.user.total_points,
                'family_id': request.user.family.id if request.user.family else None,
                'family_points': request.user.family.total_family_points if request.user.family else None
            }
            
            # Get watch history
            watch_history = list(VideoWatchHistory.objects.filter(user=request.user).values(
                'id', 'video__title', 'video__points_reward', 'points_earned', 'watched_at'
            ))
        
        videos_data = []
        for video in all_videos:
            videos_data.append({
                'id': video.id,
                'title': video.title,
                'points_reward': video.points_reward,
                'is_active': video.is_active,
                'created_at': video.created_at.isoformat()
            })
        
        return JsonResponse({
            'success': True,
            'total_videos': len(all_videos),
            'active_videos': len(active_videos),
            'videos': videos_data,
            'user_info': user_info,
            'watch_history': watch_history,
            'user_authenticated': request.user.is_authenticated
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Debug error: {str(e)}',
            'error_type': type(e).__name__
        }, status=500)

@require_http_methods(["GET"])
def test_api(request):
    """Simple test endpoint to check API connectivity"""
    return JsonResponse({
        'success': True,
        'message': 'API is working!',
        'timestamp': str(timezone.now()),
        'user_authenticated': request.user.is_authenticated,
        'user_id': request.user.id if request.user.is_authenticated else None
    })

# Quiz System Views

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def get_quiz_questions(request, video_id):
    """
    GET /api/learning/videos/{id}/quiz-questions/
    Fetch quiz questions for a specific video
    """
    try:
        video = get_object_or_404(LearningVideo, id=video_id, is_active=True)
        
        # Get up to 10 active quiz questions
        questions = QuizQuestion.objects.filter(
            video=video, 
            is_active=True
        ).order_by('order', 'id')[:10]
        
        if not questions:
            return Response({
                'success': True,
                'questions': [],
                'message': 'No quiz questions available for this video',
                'video_title': video.title
            })
        
        serializer = QuizQuestionSerializer(questions, many=True)
        return Response({
            'success': True,
            'questions': serializer.data,
            'video_title': video.title,
            'total_questions': len(serializer.data)
        })
    
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Failed to fetch quiz questions: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def submit_quiz_answers(request, video_id):
    """
    POST /api/learning/videos/{id}/quiz-submit/
    Submit quiz answers and calculate results
    """
    try:
        video = get_object_or_404(LearningVideo, id=video_id, is_active=True)
        
        serializer = QuizSubmissionSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                'success': False,
                'error': 'Invalid data format',
                'details': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        answers_data = serializer.validated_data['answers']
        
        if not answers_data:
            return Response({
                'success': False,
                'error': 'No answers provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        with transaction.atomic():
            # Check if user already has a passing quiz result for this video (70% or higher)
            existing_passed_quiz = QuizResult.objects.filter(
                user=request.user,
                video=video,
                score_percentage__gte=70.0
            ).exists()
            
            # Create or update quiz result
            quiz_result, created = QuizResult.objects.get_or_create(
                user=request.user,
                video=video,
                defaults={
                    'total_questions': len(answers_data),
                    'correct_answers': 0,
                    'score_percentage': 0.0,
                    'total_points': 0,
                }
            )
            
            # Clear existing answers if retaking quiz
            if not created:
                quiz_result.answers.all().delete()
            
            correct_count = 0
            total_points = 0
            
            # Process each answer
            for answer_data in answers_data:
                question = get_object_or_404(QuizQuestion, id=answer_data['question_id'])
                
                # Verify answer correctness
                is_correct = question.correct_answer == answer_data['selected_answer']
                points_earned = question.points_reward if is_correct else 0
                
                if is_correct:
                    correct_count += 1
                    total_points += points_earned
                
                # Save individual answer
                QuizAnswer.objects.create(
                    quiz_result=quiz_result,
                    question=question,
                    selected_answer=answer_data['selected_answer'],
                    is_correct=is_correct,
                    points_earned=points_earned
                )
            
            # Update quiz result
            quiz_result.correct_answers = correct_count
            quiz_result.score_percentage = (correct_count / len(answers_data)) * 100
            quiz_result.total_points = total_points
            quiz_result.save()
            
            # Determine if points should be awarded
            points_awarded = 0
            
            # Only award points if:
            # 1. Quiz is passed (70% or higher) AND
            # 2. User hasn't already passed this quiz before
            if quiz_result.score_percentage >= 70.0 and not existing_passed_quiz:
                points_awarded = total_points
                
                # Award points to user
                logger.debug(f"Awarding {points_awarded} points to user {request.user.username}")
                request.user.total_points += points_awarded
                request.user.save()
                logger.debug(f"User total points after award: {request.user.total_points}")
                
                # Award points to family if exists
                if request.user.family:
                    request.user.family.total_family_points += points_awarded
                    request.user.family.save()
                
                # Mark video as watched
                VideoWatchHistory.objects.get_or_create(
                    user=request.user,
                    video=video,
                    defaults={
                        'points_earned': points_awarded,
                        'watch_percentage': 100.0
                    }
                )
                
                # Create points transaction record
                PointsTransaction.objects.create(
                    user=request.user,
                    transaction_type='earned',
                    points_amount=points_awarded,
                    description=f"Completed quiz for video: {video.title}",
                    transaction_date=quiz_result.completed_at
                )
                
                # Create notification
                from accounts.models import Notification
                Notification.objects.create(
                    user=request.user,
                    type='quiz',
                    message=f'Great job! You passed the quiz for "{video.title}" with {quiz_result.score_percentage:.1f}% and earned {points_awarded} points!',
                    points=points_awarded,
                    video_title=video.title
                )
                
                logger.debug(f"Successfully awarded {points_awarded} points for quiz completion")
            
            elif quiz_result.score_percentage >= 70.0 and existing_passed_quiz:
                logger.debug(f"Quiz passed but user already completed this quiz before - no points awarded")
            
            else:
                logger.debug(f"Quiz not passed (score: {quiz_result.score_percentage}%) - no points awarded")
        
        result_serializer = QuizResultSerializer(quiz_result)
        return Response({
            'success': True,
            'result': result_serializer.data,
            'message': f'Quiz submitted successfully! Score: {quiz_result.score_percentage:.1f}%',
            'passed': quiz_result.score_percentage >= 70.0,
            'points_awarded': points_awarded  # Return actual points awarded (0 if already completed)
        })
    
    except Exception as e:
        logger.error(f"Error in submit_quiz_answers: {str(e)}")
        return Response({
            'success': False,
            'error': f'Failed to submit quiz: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def mark_video_watched(request, video_id):
    """
    POST /api/learning/videos/{id}/mark-watched/
    Mark a video as fully watched (for videos without quiz)
    """
    try:
        video = get_object_or_404(LearningVideo, id=video_id, is_active=True)
        
        # Check if video has quiz - if it does, user should complete quiz instead
        has_quiz = QuizQuestion.objects.filter(video=video, is_active=True).exists()
        if has_quiz:
            return Response({
                'success': False,
                'error': 'This video has a quiz. Please complete the quiz to earn points.',
                'has_quiz': True
            }, status=status.HTTP_400_BAD_REQUEST)
        
        watch_history, created = VideoWatchHistory.objects.get_or_create(
            user=request.user,
            video=video,
            defaults={
                'points_earned': video.points_reward,
                'watch_percentage': 100.0
            }
        )
        
        points_awarded = 0
        if created:
            # Award points to user
            points_awarded = video.points_reward
            request.user.total_points += points_awarded
            request.user.save()
            
            # Award points to family if exists
            if request.user.family:
                request.user.family.total_family_points += points_awarded
                request.user.family.save()
            
            # Create points transaction record
            PointsTransaction.objects.create(
                user=request.user,
                transaction_type='earned',
                points_amount=points_awarded,
                description=f"Watched learning video: {video.title}",
                transaction_date=watch_history.watched_at
            )
            
            # Create notification
            from accounts.models import Notification
            Notification.objects.create(
                user=request.user,
                type='learning',
                message=f'You earned {points_awarded} points for watching "{video.title}"!',
                points=points_awarded,
                video_title=video.title
            )
        
        return Response({
            'success': True,
            'message': 'Video marked as watched' if created else 'Video already watched',
            'points_earned': points_awarded,
            'new_total_points': request.user.total_points
        })
    
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Failed to mark video as watched: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def check_quiz_availability(request, video_id):
    """
    GET /api/learning/videos/{id}/has-quiz/
    Check if a video has quiz questions available
    """
    try:
        video = get_object_or_404(LearningVideo, id=video_id, is_active=True)
        
        quiz_questions = QuizQuestion.objects.filter(video=video, is_active=True)
        has_quiz = quiz_questions.exists()
        question_count = quiz_questions.count()
        
        # Check if user already completed the quiz
        quiz_completed = QuizResult.objects.filter(user=request.user, video=video).exists()
        quiz_score = None
        
        if quiz_completed:
            result = QuizResult.objects.get(user=request.user, video=video)
            quiz_score = result.score_percentage
        
        return Response({
            'success': True,
            'has_quiz': has_quiz,
            'question_count': question_count,
            'video_title': video.title,
            'quiz_completed': quiz_completed,
            'quiz_score': quiz_score
        })
    
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Failed to check quiz availability: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def get_quiz_result(request, video_id):
    """
    GET /api/learning/videos/{id}/quiz-result/
    Get user's quiz result for a specific video
    """
    try:
        video = get_object_or_404(LearningVideo, id=video_id, is_active=True)
        
        try:
            quiz_result = QuizResult.objects.get(user=request.user, video=video)
            serializer = QuizResultSerializer(quiz_result)
            
            return Response({
                'success': True,
                'result': serializer.data
            })
        except QuizResult.DoesNotExist:
            return Response({
                'success': False,
                'error': 'No quiz result found for this video',
                'completed': False
            }, status=status.HTTP_404_NOT_FOUND)
    
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Failed to fetch quiz result: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def debug_quiz_question(request, video_id, question_id):
    """
    GET /api/learning/videos/{video_id}/questions/{question_id}/debug/
    Debug a specific quiz question to check option values
    """
    try:
        video = get_object_or_404(LearningVideo, id=video_id, is_active=True)
        question = get_object_or_404(QuizQuestion, id=question_id, video=video, is_active=True)
        
        logger.debug(f"Question {question_id}: question_text='{question.question_text}', correct_answer='{question.correct_answer}'")
        
        # Test serializer output
        serializer = QuizQuestionSerializer(question)
        serialized_data = serializer.data
        
        return Response({
            'success': True,
            'raw_data': {
                'question_text': question.question_text,
                'option_a': question.option_a,
                'option_b': question.option_b,
                'option_c': question.option_c,
                'option_d': question.option_d,
                'correct_answer': question.correct_answer,
                'points_reward': question.points_reward,
                'explanation': question.explanation
            },
            'serialized_data': serialized_data,
            'options_array': [question.option_a, question.option_b, question.option_c, question.option_d],
            'non_empty_options': [opt for opt in [question.option_a, question.option_b, question.option_c, question.option_d] if opt and opt.strip()]
        })
    
    except Exception as e:
        logger.error(f"Error in debug_quiz_question: {str(e)}")
        return Response({
            'success': False,
            'error': f'Failed to debug quiz question: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ============================================================================
# EXCEL UPLOAD/DOWNLOAD FUNCTIONALITY FOR ADMIN
# ============================================================================

@admin_required
def download_quiz_template(request):
    """
    Download an Excel template for bulk quiz question upload.
    
    Template includes:
    - Header row with column names
    - Sample data row with examples
    - Formatting for clarity
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Quiz Questions Template"
    
    # Define headers
    headers = [
        'video_id', 'question_text', 'option_a', 'option_b', 
        'option_c', 'option_d', 'correct_answer', 'points_reward', 
        'explanation'
    ]
    
    # Style the header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sample data row
    sample_data = [
        1,  # video_id (get first video ID from database)
        "What type of waste can be recycled?",  # question_text
        "Plastic bottles",  # option_a
        "Food waste",  # option_b
        "Broken glass",  # option_c
        "Batteries",  # option_d
        "A",  # correct_answer
        10,  # points_reward
        "Plastic bottles are recyclable materials."  # explanation
    ]
    
    for col_num, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col_num, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Adjust column widths
    column_widths = [12, 50, 30, 30, 30, 30, 15, 15, 50]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    # Add instructions in a separate sheet
    ws_info = wb.create_sheet("Instructions")
    instructions = [
        ("Quiz Question Upload Template - Instructions", ""),
        ("", ""),
        ("Column Descriptions:", ""),
        ("video_id", "ID of the learning video (must exist in database). Optional when uploading from a specific video page."),
        ("question_text", "The quiz question text"),
        ("option_a", "First answer option"),
        ("option_b", "Second answer option"),
        ("option_c", "Third answer option"),
        ("option_d", "Fourth answer option"),
        ("correct_answer", "Correct option: A, B, C, or D"),
        ("points_reward", "Points awarded for correct answer (default: 10)"),
        ("explanation", "Explanation for the correct answer (optional)"),
        ("", ""),
        ("Important Notes:", ""),
        ("1.", "All columns are required except 'explanation'"),
        ("2.", "video_id can be left out if uploading from a specific video's quiz management page"),
        ("3.", "correct_answer must be A, B, C, or D"),
        ("4.", "Points must be a positive integer"),
        ("5.", "Order is automatically assigned based on existing questions in the database"),
        ("6.", "Delete the sample row before uploading your data"),
    ]
    
    for row_num, (col1, col2) in enumerate(instructions, 1):
        cell1 = ws_info.cell(row=row_num, column=1, value=col1)
        cell2 = ws_info.cell(row=row_num, column=2, value=col2)
        
        if row_num == 1:
            cell1.font = Font(bold=True, size=14)
        elif col1 in ["Column Descriptions:", "Important Notes:"]:
            cell1.font = Font(bold=True, size=12)
    
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 60
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create response
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=quiz_questions_template.xlsx'
    
    return response


@admin_required
@require_http_methods(["POST"])
def upload_quiz_excel(request):
    """
    Upload Excel file to bulk create quiz questions.
    
    Validates data and creates QuizQuestion objects.
    Returns detailed error report if validation fails.
    
    If video_id is provided in POST request, it will override the video_id column in Excel file,
    allowing all questions to be created for a specific video.
    """
    if 'excel_file' not in request.FILES:
        return JsonResponse({
            'success': False,
            'error': 'No file uploaded. Please select an Excel file.'
        }, status=400)
    
    excel_file = request.FILES['excel_file']
    
    # Check if video_id is provided in POST request (for per-video upload)
    override_video_id = request.POST.get('video_id')
    if override_video_id:
        try:
            override_video_id = int(override_video_id)
            # Verify video exists
            if not LearningVideo.objects.filter(id=override_video_id).exists():
                return JsonResponse({
                    'success': False,
                    'error': f'Video with ID {override_video_id} does not exist.'
                }, status=400)
        except (ValueError, TypeError):
            return JsonResponse({
                'success': False,
                'error': 'Invalid video_id provided.'
            }, status=400)
    
    # Validate file extension
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({
            'success': False,
            'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls).'
        }, status=400)
    
    try:
        # Read Excel file
        df = pd.read_excel(excel_file, sheet_name=0)
        
        # Validate required columns (video_id and order are optional)
        # order will be automatically calculated from existing questions
        if override_video_id:
            required_columns = [
                'question_text', 'option_a', 'option_b',
                'option_c', 'option_d', 'correct_answer', 'points_reward'
            ]
        else:
            required_columns = [
                'video_id', 'question_text', 'option_a', 'option_b',
                'option_c', 'option_d', 'correct_answer', 'points_reward'
            ]
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return JsonResponse({
                'success': False,
                'error': f'Missing required columns: {", ".join(missing_columns)}'
            }, status=400)
        
        # Validation errors list
        errors = []
        created_count = 0
        
        # Get the maximum order for each video to auto-increment
        video_max_orders = {}
        
        # Process each row
        with transaction.atomic():
            for index, row in df.iterrows():
                row_num = index + 2  # Excel row number (accounting for header)
                row_errors = []
                
                # Validate video_id (use override if provided)
                if override_video_id:
                    video_id = override_video_id
                else:
                    try:
                        video_id = int(row['video_id'])
                        if not LearningVideo.objects.filter(id=video_id).exists():
                            row_errors.append(f"Video ID {video_id} does not exist")
                    except (ValueError, TypeError):
                        row_errors.append(f"Invalid video_id: {row['video_id']}")
                        continue
                
                # Validate question_text
                question_text = str(row['question_text']).strip()
                if not question_text or question_text == 'nan':
                    row_errors.append("question_text is empty")
                
                # Validate options
                options = {}
                for opt in ['option_a', 'option_b', 'option_c', 'option_d']:
                    opt_value = str(row[opt]).strip()
                    if not opt_value or opt_value == 'nan':
                        row_errors.append(f"{opt} is empty")
                    options[opt] = opt_value
                
                # Validate correct_answer
                correct_answer = str(row['correct_answer']).strip().upper()
                if correct_answer not in ['A', 'B', 'C', 'D']:
                    row_errors.append(f"correct_answer must be A, B, C, or D (got: {correct_answer})")
                
                # Validate points_reward
                try:
                    points_reward = int(row['points_reward'])
                    if points_reward < 0:
                        row_errors.append("points_reward must be positive")
                except (ValueError, TypeError):
                    row_errors.append(f"Invalid points_reward: {row['points_reward']}")
                    points_reward = 10  # default
                
                # Auto-calculate order based on existing questions for this video
                # Get the max order for this video if not already cached
                if video_id not in video_max_orders:
                    max_order_obj = QuizQuestion.objects.filter(video_id=video_id).aggregate(Max('order'))
                    video_max_orders[video_id] = max_order_obj['order__max'] or 0
                
                # Increment the order for this video
                video_max_orders[video_id] += 1
                order = video_max_orders[video_id]
                
                # Get explanation (optional)
                explanation = str(row.get('explanation', '')).strip()
                if explanation == 'nan':
                    explanation = ''
                
                # If there are errors for this row, record them
                if row_errors:
                    errors.append(f"Row {row_num}: {'; '.join(row_errors)}")
                    continue
                
                # Create QuizQuestion
                try:
                    video = LearningVideo.objects.get(id=video_id)
                    QuizQuestion.objects.create(
                        video=video,
                        question_text=question_text,
                        option_a=options['option_a'],
                        option_b=options['option_b'],
                        option_c=options['option_c'],
                        option_d=options['option_d'],
                        correct_answer=correct_answer,
                        points_reward=points_reward,
                        explanation=explanation,
                        order=order,
                        is_active=True
                    )
                    created_count += 1
                except Exception as e:
                    row_errors.append(f"Database error: {str(e)}")
                    errors.append(f"Row {row_num}: {'; '.join(row_errors)}")
        
        # Return result
        if errors:
            return JsonResponse({
                'success': False,
                'created_count': created_count,
                'error_count': len(errors),
                'errors': errors[:50]  # Limit to first 50 errors
            }, status=400)
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully created {created_count} quiz questions.',
            'created_count': created_count
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Failed to process Excel file: {str(e)}'
        }, status=500)

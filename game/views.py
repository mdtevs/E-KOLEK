from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from django.views.decorators.http import require_http_methods
from cenro.admin_auth import admin_required
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from .models import Question, GameSession, WasteCategory, WasteItem, Choice
from .serializers import QuestionSerializer
from accounts.models import Users  
import logging

logger = logging.getLogger(__name__)

def home(request):
    return HttpResponse("Welcome to the Quiz App!")

@api_view(['GET'])
def debug_game_auth(request):
    """Debug endpoint to check game authentication"""
    try:
        debug_info = {
            'has_authorization_header': 'HTTP_AUTHORIZATION' in request.META,
            'auth_header': request.META.get('HTTP_AUTHORIZATION', 'Not present'),
            'user_authenticated': request.user.is_authenticated,
            'user_info': str(request.user),
            'request_method': request.method,
            'request_path': request.path,
        }
        return Response({
            'success': True,
            'debug': debug_info
        })
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        })

@api_view(['GET'])
def test_questions_no_auth(request):
    """Temporary endpoint to test questions without authentication"""
    try:
        questions = Question.objects.all()
        data = []
        
        for question in questions[:2]:  # Just return 2 questions for testing
            question_data = {
                'id': question.id,
                'text': question.text,
                'points': question.points,
                'choices': []
            }
            
            for choice in question.choices.all():
                choice_data = {
                    'id': choice.id,
                    'text': choice.text,
                    'is_correct': choice.is_correct
                }
                question_data['choices'].append(choice_data)
            
            data.append(question_data)
        
        return Response({
            'success': True,
            'questions': data,
            'count': len(data)
        })
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        })

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def questions_view(request):
    """Get questions with points included - for quiz game"""
    try:
        # Check if user is authenticated
        if not request.user.is_authenticated:
            return Response({
                'error': 'Authentication required',
                'detail': 'User is not authenticated'
            }, status=401)
        
        questions = Question.objects.all()
        data = []
        
        for question in questions:
            question_data = {
                'id': question.id,
                'text': question.text,
                'points': question.points,  # Include points in API response
                'choices': []
            }
            
            for choice in question.choices.all():
                choice_data = {
                    'id': choice.id,
                    'text': choice.text,
                    'is_correct': choice.is_correct
                }
                question_data['choices'].append(choice_data)
            
            data.append(question_data)
        
        logger.debug(f"Questions view - Returning {len(data)} questions")
        return Response(data)
    except Exception as e:
        logger.error(f"Error in questions_view: {str(e)}")
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def get_game_data(request):
    """Get waste categories and items for the game"""
    try:
        logger.debug(f"Game data view - User: {request.user}, Auth: {request.user.is_authenticated}")
        logger.debug(f"Game data view - Auth header: {request.META.get('HTTP_AUTHORIZATION', 'Not present')}")
        
        categories = WasteCategory.objects.all()
        items = WasteItem.objects.filter(is_active=True)
        
        categories_data = []
        for category in categories:
            categories_data.append({
                'id': str(category.id),
                'name': category.name,
                'color_hex': category.color_hex,
                'icon_name': category.icon_name,
                'description': category.description
            })
        
        items_data = []
        for item in items:
            items_data.append({
                'id': str(item.id),
                'name': item.name,
                'emoji': item.emoji,
                'category_id': str(item.category.id),
                'category_name': item.category.name,
                'points': item.points,
                'difficulty_level': item.difficulty_level
            })
        
        logger.debug(f"Game data view - Returning {len(categories_data)} categories, {len(items_data)} items")
        return Response({
            'success': True,
            'categories': categories_data,
            'items': items_data
        })
    except Exception as e:
        logger.error(f"Error in get_game_data: {str(e)}")
        return Response({'success': False, 'error': str(e)}, status=500)

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def save_game_session(request):
    """Save completed game session"""
    try:
        data = request.data
        
        with transaction.atomic():
            # Get the authenticated user (should be a Users instance)
            user = request.user
            logger.debug(f"Game session for user: {user.full_name} ({user.username})")
            
            session = GameSession.objects.create(
                user=user,
                score=data.get('score', 0),
                correct_answers=data.get('correct_answers', 0),
                wrong_answers=data.get('wrong_answers', 0),
                accuracy=data.get('accuracy', 0.0),
                duration_seconds=data.get('duration_seconds', 0)
            )
            
            # Update user's total points
            points_earned = data.get('score', 0)
            user.total_points += points_earned
            user.save()
            
            logger.debug(f"Added {points_earned} points. New total: {user.total_points}")
            
            # Create notification for game completion (only if points were earned)
            if points_earned > 0:
                from accounts.models import Notification
                Notification.objects.create(
                    user=user,
                    type='game',
                    message=f'You earned {points_earned} points from playing the waste sorting game!',
                    points=points_earned,
                    game_score=data.get('score', 0)
                )
            
            # Update family points as well (this will be handled by the user's save method)
            # The custom save() method in Users model will automatically update family points
        
        return Response({
            'success': True,
            'session_id': str(session.id),
            'new_total_points': user.total_points,
            'points_earned': points_earned
        })
    except Exception as e:
        logger.error(f"Error in save_game_session: {str(e)}")
        return Response({'success': False, 'error': str(e)}, status=400)

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def get_game_leaderboard(request):
    """Get top game scores aggregated by user"""
    try:
        # Debug authentication
        auth_header = request.META.get('HTTP_AUTHORIZATION', 'Not provided')
        logger.debug(f"Auth header in game leaderboard: {auth_header}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        
        from django.db.models import Sum, Count, Avg, Max
        
        # Aggregate all sessions by user to get total scores
        user_stats = GameSession.objects.select_related('user', 'user__family').values(
            'user__username',
            'user__full_name', 
            'user__family__family_name'
        ).annotate(
            total_score=Sum('score'),
            total_games=Count('id'),
            avg_accuracy=Avg('accuracy'),
            best_score=Max('score')
        ).order_by('-total_score')[:50]  # Get top 50 users instead of top 10 sessions
        
        leaderboard = []
        for stats in user_stats:
            leaderboard.append({
                'username': stats['user__username'],
                'full_name': stats['user__full_name'],
                'family_name': stats['user__family__family_name'],
                'score': stats['total_score'],  # This is now total score across all games
                'total_games': stats['total_games'],
                'accuracy': round(stats['avg_accuracy'], 2) if stats['avg_accuracy'] else 0,
                'best_score': stats['best_score']
            })
        
        return Response({
            'success': True,
            'leaderboard': leaderboard
        })
    except Exception as e:
        logger.error(f"Error in get_game_leaderboard: {str(e)}")
        return Response({'success': False, 'error': str(e)}, status=400)

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def get_user_game_stats(request):
    """Get current user's game statistics"""
    try:
        user = request.user
        
        # Get user's game sessions
        user_sessions = GameSession.objects.filter(user=user).order_by('-completed_at')
        
        # Calculate statistics
        total_games = user_sessions.count()
        total_score = sum(session.score for session in user_sessions)
        avg_accuracy = sum(session.accuracy for session in user_sessions) / total_games if total_games > 0 else 0
        best_score = max(session.score for session in user_sessions) if total_games > 0 else 0
        
        # Get recent sessions
        recent_sessions = []
        for session in user_sessions[:5]:  # Last 5 games
            recent_sessions.append({
                'score': session.score,
                'accuracy': session.accuracy,
                'completed_at': session.completed_at.isoformat()
            })
        
        return Response({
            'success': True,
            'user_info': {
                'username': user.username,
                'full_name': user.full_name,
                'family_name': user.family.family_name,
                'total_points': user.total_points
            },
            'game_stats': {
                'total_games': total_games,
                'total_score': total_score,
                'average_accuracy': round(avg_accuracy, 2),
                'best_score': best_score
            },
            'recent_sessions': recent_sessions
        })
    except Exception as e:
        logger.error(f"Error in get_user_game_stats: {str(e)}")
        return Response({'success': False, 'error': str(e)}, status=400)


# ============================================================================
# EXCEL UPLOAD/DOWNLOAD FUNCTIONALITY FOR ADMIN
# ============================================================================

@admin_required
def download_game_template(request):
    """
    Download an Excel template for bulk game question upload.
    
    Template includes:
    - Header row with column names
    - Sample data row with examples
    - Formatting for clarity
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Game Questions Template"
    
    # Define headers
    headers = [
        'question_text', 'points', 'choice_1', 'choice_2', 
        'choice_3', 'choice_4', 'correct_choice_number'
    ]
    
    # Style the header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sample data row
    sample_data = [
        "What bin should plastic bottles go in?",  # question_text
        10,  # points
        "Recyclable Bin",  # choice_1
        "Organic Bin",  # choice_2
        "Hazardous Bin",  # choice_3
        "General Waste",  # choice_4
        1   # correct_choice_number (1-4, indicating choice_1 is correct)
    ]
    
    for col_num, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col_num, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Adjust column widths
    column_widths = [50, 12, 30, 30, 30, 30, 22]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    # Add instructions in a separate sheet
    ws_info = wb.create_sheet("Instructions")
    instructions = [
        ("Game Question Upload Template - Instructions", ""),
        ("", ""),
        ("Column Descriptions:", ""),
        ("question_text", "The game question text"),
        ("points", "Points awarded for correct answer (default: 1)"),
        ("choice_1", "First answer choice"),
        ("choice_2", "Second answer choice"),
        ("choice_3", "Third answer choice (optional)"),
        ("choice_4", "Fourth answer choice (optional)"),
        ("correct_choice_number", "Number of the correct choice (1, 2, 3, or 4)"),
        ("", ""),
        ("Important Notes:", ""),
        ("1.", "All columns are required except choice_3 and choice_4"),
        ("2.", "At least 2 choices (choice_1 and choice_2) must be provided"),
        ("3.", "correct_choice_number must be 1, 2, 3, or 4"),
        ("4.", "The correct_choice_number must correspond to a non-empty choice"),
        ("5.", "Points must be a positive integer"),
        ("6.", "Leave choice_3 or choice_4 empty if not needed"),
        ("7.", "Delete the sample row before uploading your data"),
    ]
    
    for row_num, (col1, col2) in enumerate(instructions, 1):
        cell1 = ws_info.cell(row=row_num, column=1, value=col1)
        cell2 = ws_info.cell(row=row_num, column=2, value=col2)
        
        if row_num == 1:
            cell1.font = Font(bold=True, size=14)
        elif col1 in ["Column Descriptions:", "Important Notes:"]:
            cell1.font = Font(bold=True, size=12)
    
    ws_info.column_dimensions['A'].width = 25
    ws_info.column_dimensions['B'].width = 60
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create response
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=game_questions_template.xlsx'
    
    return response


@admin_required
@require_http_methods(["POST"])
def upload_game_excel(request):
    """
    Upload Excel file to bulk create game questions.
    
    Validates data and creates Question and Choice objects.
    Returns detailed error report if validation fails.
    """
    if 'excel_file' not in request.FILES:
        return JsonResponse({
            'success': False,
            'error': 'No file uploaded. Please select an Excel file.'
        }, status=400)
    
    excel_file = request.FILES['excel_file']
    
    # Validate file extension
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({
            'success': False,
            'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls).'
        }, status=400)
    
    try:
        # Read Excel file
        df = pd.read_excel(excel_file, sheet_name=0)
        
        # Validate required columns
        required_columns = [
            'question_text', 'points', 'choice_1', 'choice_2', 'correct_choice_number'
        ]
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return JsonResponse({
                'success': False,
                'error': f'Missing required columns: {", ".join(missing_columns)}'
            }, status=400)
        
        # Validation errors list
        errors = []
        created_count = 0
        
        # Process each row
        with transaction.atomic():
            for index, row in df.iterrows():
                row_num = index + 2  # Excel row number (accounting for header)
                row_errors = []
                
                # Validate question_text
                question_text = str(row['question_text']).strip()
                if not question_text or question_text == 'nan':
                    row_errors.append("question_text is empty")
                
                # Validate points
                try:
                    points = int(row['points'])
                    if points < 0:
                        row_errors.append("points must be positive")
                except (ValueError, TypeError):
                    row_errors.append(f"Invalid points: {row['points']}")
                    points = 1  # default
                
                # Collect choices
                choices = []
                for i in range(1, 5):
                    choice_col = f'choice_{i}'
                    if choice_col in df.columns:
                        choice_text = str(row.get(choice_col, '')).strip()
                        if choice_text and choice_text != 'nan':
                            choices.append(choice_text)
                        elif i <= 2:  # First two choices are required
                            row_errors.append(f"{choice_col} is empty (required)")
                
                # Validate we have at least 2 choices
                if len(choices) < 2:
                    row_errors.append("At least 2 choices are required")
                
                # Validate correct_choice_number
                try:
                    correct_choice_number = int(row['correct_choice_number'])
                    if correct_choice_number < 1 or correct_choice_number > len(choices):
                        row_errors.append(
                            f"correct_choice_number must be between 1 and {len(choices)} "
                            f"(got: {correct_choice_number})"
                        )
                except (ValueError, TypeError):
                    row_errors.append(f"Invalid correct_choice_number: {row['correct_choice_number']}")
                    correct_choice_number = 1  # default
                
                # If there are errors for this row, record them
                if row_errors:
                    errors.append(f"Row {row_num}: {'; '.join(row_errors)}")
                    continue
                
                # Create Question and Choices
                try:
                    question = Question.objects.create(
                        text=question_text,
                        points=points
                    )
                    
                    # Create Choice objects
                    for idx, choice_text in enumerate(choices, 1):
                        is_correct = (idx == correct_choice_number)
                        Choice.objects.create(
                            question=question,
                            text=choice_text,
                            is_correct=is_correct
                        )
                    
                    created_count += 1
                except Exception as e:
                    row_errors.append(f"Database error: {str(e)}")
                    errors.append(f"Row {row_num}: {'; '.join(row_errors)}")
                    # Rollback will happen automatically due to transaction.atomic()
        
        # Return result
        if errors:
            return JsonResponse({
                'success': False,
                'created_count': created_count,
                'error_count': len(errors),
                'errors': errors[:50]  # Limit to first 50 errors
            }, status=400)
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully created {created_count} game questions.',
            'created_count': created_count
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Failed to process Excel file: {str(e)}'
        }, status=500)


@admin_required
def download_category_template(request):
    """
    Download an Excel template for bulk waste category upload.
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Waste Categories"
    
    # Define headers
    headers = ['name', 'color_hex', 'icon_name', 'description']
    
    # Style the header row
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sample data rows
    sample_data = [
        ["Recyclable", "#4CAF50", "recycling", "Items that can be recycled"],
        ["Organic", "#8BC34A", "nature", "Biodegradable waste"],
        ["Hazardous", "#F44336", "warning", "Dangerous materials"]
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Adjust column widths
    column_widths = [20, 15, 20, 50]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    # Add instructions sheet
    ws_info = wb.create_sheet("Instructions")
    instructions = [
        ("Waste Category Upload Template - Instructions", ""),
        ("", ""),
        ("Column Descriptions:", ""),
        ("name", "Category name (must be unique, e.g., 'Recyclable', 'Organic')"),
        ("color_hex", "Hex color code (e.g., '#4CAF50')"),
        ("icon_name", "Material icon name (e.g., 'recycling', 'delete', 'nature')"),
        ("description", "Description of the category (optional)"),
        ("", ""),
        ("Important Notes:", ""),
        ("1.", "All columns are required except 'description'"),
        ("2.", "Category names must be unique"),
        ("3.", "color_hex must be a valid hex code starting with #"),
        ("4.", "icon_name should be a valid Material icon name"),
        ("5.", "Delete the sample rows before uploading your data"),
    ]
    
    for row_num, (col1, col2) in enumerate(instructions, 1):
        cell1 = ws_info.cell(row=row_num, column=1, value=col1)
        cell2 = ws_info.cell(row=row_num, column=2, value=col2)
        
        if row_num == 1:
            cell1.font = Font(bold=True, size=14)
        elif col1 in ["Column Descriptions:", "Important Notes:"]:
            cell1.font = Font(bold=True, size=12)
    
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 60
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create response
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=waste_categories_template.xlsx'
    
    return response


@admin_required
@require_http_methods(["POST"])
def upload_category_excel(request):
    """
    Upload Excel file to bulk create waste categories.
    """
    if 'excel_file' not in request.FILES:
        return JsonResponse({
            'success': False,
            'error': 'No file uploaded. Please select an Excel file.'
        }, status=400)
    
    excel_file = request.FILES['excel_file']
    
    # Validate file extension
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({
            'success': False,
            'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls).'
        }, status=400)
    
    try:
        # Read Excel file
        df = pd.read_excel(excel_file, sheet_name=0)
        
        # Validate required columns
        required_columns = ['name', 'color_hex', 'icon_name']
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return JsonResponse({
                'success': False,
                'error': f'Missing required columns: {", ".join(missing_columns)}'
            }, status=400)
        
        # Validation errors list
        errors = []
        created_count = 0
        
        # Process each row
        with transaction.atomic():
            for index, row in df.iterrows():
                row_num = index + 2  # Excel row number (accounting for header)
                row_errors = []
                
                # Validate name
                name = str(row['name']).strip()
                if not name or name == 'nan':
                    row_errors.append("name is empty")
                elif WasteCategory.objects.filter(name=name).exists():
                    row_errors.append(f"Category '{name}' already exists")
                
                # Validate color_hex
                color_hex = str(row['color_hex']).strip()
                if not color_hex or color_hex == 'nan':
                    row_errors.append("color_hex is empty")
                elif not color_hex.startswith('#') or len(color_hex) != 7:
                    row_errors.append(f"Invalid color_hex format: {color_hex} (should be #RRGGBB)")
                
                # Validate icon_name
                icon_name = str(row['icon_name']).strip()
                if not icon_name or icon_name == 'nan':
                    row_errors.append("icon_name is empty")
                
                # Get description (optional)
                description = str(row.get('description', '')).strip()
                if description == 'nan':
                    description = ''
                
                # If there are errors for this row, record them
                if row_errors:
                    errors.append(f"Row {row_num}: {'; '.join(row_errors)}")
                    continue
                
                # Create WasteCategory
                try:
                    WasteCategory.objects.create(
                        name=name,
                        color_hex=color_hex,
                        icon_name=icon_name,
                        description=description
                    )
                    created_count += 1
                except Exception as e:
                    row_errors.append(f"Database error: {str(e)}")
                    errors.append(f"Row {row_num}: {'; '.join(row_errors)}")
        
        # Return result
        if errors:
            return JsonResponse({
                'success': False,
                'created_count': created_count,
                'error_count': len(errors),
                'errors': errors[:50]  # Limit to first 50 errors
            }, status=400)
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully created {created_count} waste categories.',
            'created_count': created_count
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Failed to process Excel file: {str(e)}'
        }, status=500)


@admin_required
def download_item_template(request):
    """
    Download an Excel template for bulk waste item upload.
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Waste Items"
    
    # Define headers
    headers = ['name', 'emoji', 'category_name', 'points', 'difficulty_level', 'is_active']
    
    # Style the header row
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sample data rows
    sample_data = [
        ["Plastic Bottle", "🥤", "Recyclable", 10, "easy", "TRUE"],
        ["Banana Peel", "🍌", "Organic", 10, "easy", "TRUE"],
        ["Battery", "🔋", "Hazardous", 20, "hard", "TRUE"]
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Adjust column widths
    column_widths = [25, 10, 20, 12, 18, 12]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    # Add instructions sheet
    ws_info = wb.create_sheet("Instructions")
    instructions = [
        ("Waste Item Upload Template - Instructions", ""),
        ("", ""),
        ("Column Descriptions:", ""),
        ("name", "Item name (e.g., 'Plastic Bottle', 'Banana Peel')"),
        ("emoji", "Emoji representation (e.g., '🥤', '🍌', '🔋') - optional"),
        ("category_name", "Category name that must already exist in database"),
        ("points", "Points awarded for correct sorting (default: 10)"),
        ("difficulty_level", "Difficulty: 'easy', 'medium', or 'hard'"),
        ("is_active", "TRUE or FALSE - whether item is active in game"),
        ("", ""),
        ("Important Notes:", ""),
        ("1.", "All columns are required except 'emoji'"),
        ("2.", "category_name must match an existing category exactly"),
        ("3.", "difficulty_level must be: easy, medium, or hard"),
        ("4.", "is_active must be: TRUE or FALSE"),
        ("5.", "points must be a positive integer"),
        ("6.", "Delete the sample rows before uploading your data"),
    ]
    
    for row_num, (col1, col2) in enumerate(instructions, 1):
        cell1 = ws_info.cell(row=row_num, column=1, value=col1)
        cell2 = ws_info.cell(row=row_num, column=2, value=col2)
        
        if row_num == 1:
            cell1.font = Font(bold=True, size=14)
        elif col1 in ["Column Descriptions:", "Important Notes:"]:
            cell1.font = Font(bold=True, size=12)
    
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 60
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create response
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=waste_items_template.xlsx'
    
    return response


@admin_required
@require_http_methods(["POST"])
def upload_item_excel(request):
    """
    Upload Excel file to bulk create waste items.
    """
    if 'excel_file' not in request.FILES:
        return JsonResponse({
            'success': False,
            'error': 'No file uploaded. Please select an Excel file.'
        }, status=400)
    
    excel_file = request.FILES['excel_file']
    
    # Validate file extension
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({
            'success': False,
            'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls).'
        }, status=400)
    
    try:
        # Read Excel file
        df = pd.read_excel(excel_file, sheet_name=0)
        
        # Validate required columns
        required_columns = ['name', 'category_name', 'points', 'difficulty_level', 'is_active']
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return JsonResponse({
                'success': False,
                'error': f'Missing required columns: {", ".join(missing_columns)}'
            }, status=400)
        
        # Validation errors list
        errors = []
        created_count = 0
        
        # Process each row
        with transaction.atomic():
            for index, row in df.iterrows():
                row_num = index + 2  # Excel row number (accounting for header)
                row_errors = []
                
                # Validate name
                name = str(row['name']).strip()
                if not name or name == 'nan':
                    row_errors.append("name is empty")
                
                # Get emoji (optional)
                emoji = str(row.get('emoji', '')).strip()
                if emoji == 'nan':
                    emoji = ''
                
                # Validate category_name
                category_name = str(row['category_name']).strip()
                if not category_name or category_name == 'nan':
                    row_errors.append("category_name is empty")
                else:
                    try:
                        category = WasteCategory.objects.get(name=category_name)
                    except WasteCategory.DoesNotExist:
                        row_errors.append(f"Category '{category_name}' does not exist")
                        category = None
                
                # Validate points
                try:
                    points = int(row['points'])
                    if points < 0:
                        row_errors.append("points must be positive")
                except (ValueError, TypeError):
                    row_errors.append(f"Invalid points: {row['points']}")
                    points = 10  # default
                
                # Validate difficulty_level
                difficulty_level = str(row['difficulty_level']).strip().lower()
                if difficulty_level not in ['easy', 'medium', 'hard']:
                    row_errors.append(f"difficulty_level must be 'easy', 'medium', or 'hard' (got: {difficulty_level})")
                    difficulty_level = 'easy'  # default
                
                # Validate is_active
                is_active_str = str(row['is_active']).strip().upper()
                if is_active_str in ['TRUE', '1', 'YES', 'Y']:
                    is_active = True
                elif is_active_str in ['FALSE', '0', 'NO', 'N']:
                    is_active = False
                else:
                    row_errors.append(f"is_active must be TRUE or FALSE (got: {row['is_active']})")
                    is_active = True  # default
                
                # If there are errors for this row, record them
                if row_errors:
                    errors.append(f"Row {row_num}: {'; '.join(row_errors)}")
                    continue
                
                # Create WasteItem
                try:
                    WasteItem.objects.create(
                        name=name,
                        emoji=emoji,
                        category=category,
                        points=points,
                        difficulty_level=difficulty_level,
                        is_active=is_active
                    )
                    created_count += 1
                except Exception as e:
                    row_errors.append(f"Database error: {str(e)}")
                    errors.append(f"Row {row_num}: {'; '.join(row_errors)}")
        
        # Return result
        if errors:
            return JsonResponse({
                'success': False,
                'created_count': created_count,
                'error_count': len(errors),
                'errors': errors[:50]  # Limit to first 50 errors
            }, status=400)
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully created {created_count} waste items.',
            'created_count': created_count
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Failed to process Excel file: {str(e)}'
        }, status=500)


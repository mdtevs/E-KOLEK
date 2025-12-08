"""
Waste Analytics Views for CENRO Admin
Provides comprehensive analytics, Excel import/export, and PDF reporting
"""

from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse, FileResponse
from django.db.models import Sum, Count, Avg, Q, F
from django.db.models.functions import TruncDate, TruncMonth, TruncWeek
from django.utils import timezone
from django.contrib import messages
from datetime import datetime, timedelta
from decimal import Decimal
import json
import io
import os

# Excel and PDF libraries
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.pdfgen import canvas
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import matplotlib
    matplotlib.use('Agg')  # Use non-interactive backend
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    from matplotlib.backends.backend_pdf import PdfPages
    MATPLOTLIB_AVAILABLE = True
    print(f"Matplotlib {matplotlib.__version__} loaded successfully")
except ImportError as e:
    MATPLOTLIB_AVAILABLE = False
    print(f"Matplotlib import failed: {e}")
    print(f"Make sure you're running Django with the virtual environment Python:")
    print(f"Use: env\\Scripts\\python.exe manage.py runserver 0.0.0.0:8000")
except Exception as e:
    MATPLOTLIB_AVAILABLE = False
    print(f"Matplotlib error: {e}")

from accounts.models import (
    WasteTransaction, WasteType, Barangay, Users, Family
)
from cenro.models import AdminUser
from .admin_auth import admin_required, permission_required
from cenro.admin_utils import log_admin_action

import logging
import sys
logger = logging.getLogger(__name__)

# Print Python path for debugging
print("=" * 80)
print(f"Python executable: {sys.executable}")
print(f"Python version: {sys.version}")
print("=" * 80)


def _fix_missing_barangays():
    """
    Helper function to populate missing barangays in existing transactions
    This runs once per dashboard load to ensure data consistency
    """
    try:
        # Find transactions without barangay but user has family with barangay
        transactions_to_fix = WasteTransaction.objects.filter(
            barangay__isnull=True,
            user__family__barangay__isnull=False
        ).select_related('user__family__barangay')
        
        if transactions_to_fix.exists():
            for transaction in transactions_to_fix:
                transaction.barangay = transaction.user.family.barangay
                transaction.save(update_fields=['barangay'])
            logger.info(f"Fixed {transactions_to_fix.count()} transactions with missing barangays")
    except Exception as e:
        logger.error(f"Error fixing missing barangays: {str(e)}")


def _calculate_year_comparison(year1, year2, barangay_id=None, waste_type_id=None):
    """
    Calculate year-over-year comparison metrics
    Returns comparison data including growth percentages
    """
    try:
        # Build querysets for both years
        year1_transactions = WasteTransaction.objects.filter(
            transaction_date__year=year1
        )
        year2_transactions = WasteTransaction.objects.filter(
            transaction_date__year=year2
        )
        
        # Apply additional filters if provided
        if barangay_id:
            year1_transactions = year1_transactions.filter(barangay_id=barangay_id)
            year2_transactions = year2_transactions.filter(barangay_id=barangay_id)
        if waste_type_id:
            year1_transactions = year1_transactions.filter(waste_type_id=waste_type_id)
            year2_transactions = year2_transactions.filter(waste_type_id=waste_type_id)
        
        # Calculate metrics for year 1
        year1_stats = year1_transactions.aggregate(
            total_weight=Sum('waste_kg'),
            total_points=Sum('total_points'),
            total_count=Count('id'),
            unique_users=Count('user', distinct=True)
        )
        
        # Calculate metrics for year 2
        year2_stats = year2_transactions.aggregate(
            total_weight=Sum('waste_kg'),
            total_points=Sum('total_points'),
            total_count=Count('id'),
            unique_users=Count('user', distinct=True)
        )
        
        # Extract values with defaults
        year1_weight = float(year1_stats['total_weight'] or 0)
        year2_weight = float(year2_stats['total_weight'] or 0)
        year1_points = float(year1_stats['total_points'] or 0)
        year2_points = float(year2_stats['total_points'] or 0)
        year1_count = year1_stats['total_count'] or 0
        year2_count = year2_stats['total_count'] or 0
        year1_users = year1_stats['unique_users'] or 0
        year2_users = year2_stats['unique_users'] or 0
        
        # Calculate percentage changes
        weight_change = ((year2_weight - year1_weight) / year1_weight * 100) if year1_weight > 0 else 0
        points_change = ((year2_points - year1_points) / year1_points * 100) if year1_points > 0 else 0
        count_change = ((year2_count - year1_count) / year1_count * 100) if year1_count > 0 else 0
        users_change = ((year2_users - year1_users) / year1_users * 100) if year1_users > 0 else 0
        
        return {
            'year1': year1,
            'year2': year2,
            'year1_weight': year1_weight,
            'year2_weight': year2_weight,
            'year1_points': year1_points,
            'year2_points': year2_points,
            'year1_count': year1_count,
            'year2_count': year2_count,
            'year1_users': year1_users,
            'year2_users': year2_users,
            'weight_change': weight_change,
            'points_change': points_change,
            'count_change': count_change,
            'users_change': users_change,
        }
    except Exception as e:
        logger.error(f"Error calculating year comparison: {e}")
        return None


@admin_required
def waste_analytics_dashboard(request):
    """
    Main analytics dashboard for waste collection data
    Shows comprehensive statistics, charts, and filters with year-based tracking
    """
    # Fix any transactions with missing barangays
    _fix_missing_barangays()
    
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    barangay_id = request.GET.get('barangay')
    waste_type_id = request.GET.get('waste_type')
    year_filter = request.GET.get('year_filter', '')
    compare_year1 = request.GET.get('compare_year1')
    compare_year2 = request.GET.get('compare_year2')
    
    # Base queryset
    transactions = WasteTransaction.objects.select_related(
        'user', 'waste_type', 'barangay', 'processed_by'
    ).all()
    
    # Handle year-based filtering
    current_year = timezone.now().year
    year_comparison = None
    
    # Default to current year if no year filter is specified
    if not year_filter or year_filter == 'current':
        start_date = f'{current_year}-01-01'
        end_date = f'{current_year}-12-31'
        year_filter = 'current'  # Set for template display
    elif year_filter == 'compare':
        # Handle custom year comparison if years are specified
        if compare_year1 and compare_year2:
            try:
                year1 = int(compare_year1)
                year2 = int(compare_year2)
                year_comparison = _calculate_year_comparison(year1, year2, barangay_id, waste_type_id)
            except (ValueError, TypeError):
                # If invalid years, fall back to current vs previous year
                year_comparison = _calculate_year_comparison(current_year, current_year - 1, barangay_id, waste_type_id)
        else:
            # Default comparison: current year vs previous year
            year_comparison = _calculate_year_comparison(current_year, current_year - 1, barangay_id, waste_type_id)
        # Don't filter transactions for compare mode - show all data
        start_date = None
        end_date = None
    elif year_filter in ['2024', '2023', '2022', '2021', '2020']:
        start_date = f'{year_filter}-01-01'
        end_date = f'{year_filter}-12-31'
    
    # Apply filters
    if start_date:
        transactions = transactions.filter(transaction_date__gte=start_date)
    if end_date:
        transactions = transactions.filter(transaction_date__lte=end_date)
    if barangay_id:
        transactions = transactions.filter(barangay_id=barangay_id)
    if waste_type_id:
        transactions = transactions.filter(waste_type_id=waste_type_id)
    
    # Calculate summary statistics
    total_transactions = transactions.count()
    total_weight = transactions.aggregate(total=Sum('waste_kg'))['total'] or 0
    total_points = transactions.aggregate(total=Sum('total_points'))['total'] or 0
    avg_weight = transactions.aggregate(avg=Avg('waste_kg'))['avg'] or 0
    
    # Waste type breakdown
    waste_type_stats = transactions.values(
        'waste_type__name'
    ).annotate(
        total_weight=Sum('waste_kg'),
        total_points=Sum('total_points'),
        count=Count('id')
    ).order_by('-total_weight')
    
    # Barangay breakdown
    barangay_stats = transactions.values(
        'barangay__name'
    ).annotate(
        total_weight=Sum('waste_kg'),
        total_points=Sum('total_points'),
        count=Count('id')
    ).order_by('-total_weight')
    
    # Calculate insights
    unique_users = transactions.values('user').distinct().count()
    unique_barangays = transactions.filter(barangay__isnull=False).values('barangay').distinct().count()
    
    # Find most popular waste type
    most_popular_waste = waste_type_stats.first() if waste_type_stats else None
    
    # Find most active barangay
    most_active_barangay = barangay_stats.first() if barangay_stats else None
    
    # Calculate environmental impact (example: 1kg waste = 0.5kg CO2 saved)
    co2_saved = total_weight * 0.5
    trees_equivalent = total_weight * 0.02  # 1kg waste = 0.02 trees planted equivalent
    
    # Top collectors (users)
    top_collectors = transactions.values(
        'user__full_name', 'user__family__family_code'
    ).annotate(
        total_weight=Sum('waste_kg'),
        total_points=Sum('total_points'),
        count=Count('id')
    ).order_by('-total_weight')[:10]
    
    # Recent transactions
    recent_transactions = transactions.order_by('-created_at')[:20]
    
    # Get all barangays and waste types for filters
    all_barangays = Barangay.objects.all().order_by('name')
    all_waste_types = WasteType.objects.all().order_by('name')
    
    # Prepare chart data for waste type distribution (doughnut)
    waste_type_chart_data = {
        'labels': [item['waste_type__name'] for item in waste_type_stats],
        'weights': [float(item['total_weight']) for item in waste_type_stats],
        'points': [float(item['total_points']) for item in waste_type_stats],
        'counts': [item['count'] for item in waste_type_stats]
    }
    
    # Prepare chart data for barangay distribution (horizontal bar)
    barangay_chart_data = {
        'labels': [item['barangay__name'] for item in barangay_stats if item['barangay__name']],
        'weights': [float(item['total_weight']) for item in barangay_stats if item['barangay__name']],
        'counts': [item['count'] for item in barangay_stats if item['barangay__name']]
    }
    
    # Prepare chart data for waste type comparison (vertical bar chart)
    waste_type_bar_data = {
        'labels': [item['waste_type__name'] for item in waste_type_stats],
        'weights': [float(item['total_weight']) for item in waste_type_stats],
        'transactions': [item['count'] for item in waste_type_stats],
        'points': [float(item['total_points']) for item in waste_type_stats]
    }
    
    # ==================== MONTHLY WASTE TYPE TRACKING ====================
    # Aggregate transactions by month and waste type for trend analysis
    monthly_waste_data = transactions.annotate(
        month=TruncMonth('transaction_date')
    ).values('month', 'waste_type__name').annotate(
        total_weight=Sum('waste_kg'),
        transaction_count=Count('id')
    ).order_by('month', 'waste_type__name')
    
    # Organize data for multi-line chart (each waste type gets its own line)
    monthly_chart_data = {}
    all_months = []
    waste_type_names = [wt['waste_type__name'] for wt in waste_type_stats]
    
    # Extract all unique months
    for item in monthly_waste_data:
        month_str = item['month'].strftime('%Y-%m') if item['month'] else 'Unknown'
        if month_str not in all_months:
            all_months.append(month_str)
    
    # Initialize data structure for each waste type
    for waste_name in waste_type_names:
        monthly_chart_data[waste_name] = {
            'weights': [0] * len(all_months),
            'counts': [0] * len(all_months)
        }
    
    # Populate the data
    for item in monthly_waste_data:
        if item['month']:
            month_str = item['month'].strftime('%Y-%m')
            waste_name = item['waste_type__name']
            month_index = all_months.index(month_str)
            
            if waste_name in monthly_chart_data:
                monthly_chart_data[waste_name]['weights'][month_index] = float(item['total_weight'])
                monthly_chart_data[waste_name]['counts'][month_index] = item['transaction_count']
    
    # Format months for display (e.g., "Jan 2025")
    formatted_months = []
    for month_str in all_months:
        try:
            date_obj = datetime.strptime(month_str, '%Y-%m')
            formatted_months.append(date_obj.strftime('%b %Y'))
        except:
            formatted_months.append(month_str)
    
    # Calculate monthly insights
    monthly_insights = {
        'total_months': len(all_months),
        'peak_month': None,
        'peak_month_weight': 0,
        'trend': 'stable',
        'growth_rate': 0,
        'most_consistent_type': None,
        'most_variable_type': None,
    }
    
    # Find peak collection month
    monthly_totals = {}
    for month_idx, month in enumerate(all_months):
        total = sum(monthly_chart_data[wt]['weights'][month_idx] for wt in waste_type_names if wt in monthly_chart_data)
        monthly_totals[month] = total
        if total > monthly_insights['peak_month_weight']:
            monthly_insights['peak_month_weight'] = total
            monthly_insights['peak_month'] = formatted_months[month_idx]
    
    # Calculate trend (compare first half vs second half)
    if len(all_months) >= 2:
        mid_point = len(all_months) // 2
        first_half_avg = sum(list(monthly_totals.values())[:mid_point]) / mid_point if mid_point > 0 else 0
        second_half_avg = sum(list(monthly_totals.values())[mid_point:]) / (len(all_months) - mid_point) if (len(all_months) - mid_point) > 0 else 0
        
        if second_half_avg > first_half_avg * 1.1:
            monthly_insights['trend'] = 'increasing'
            monthly_insights['growth_rate'] = ((second_half_avg - first_half_avg) / first_half_avg * 100) if first_half_avg > 0 else 0
        elif second_half_avg < first_half_avg * 0.9:
            monthly_insights['trend'] = 'decreasing'
            monthly_insights['growth_rate'] = ((first_half_avg - second_half_avg) / first_half_avg * 100) if first_half_avg > 0 else 0
        else:
            monthly_insights['trend'] = 'stable'
    
    # Find most consistent and variable waste types
    if waste_type_names:
        variability_scores = {}
        for waste_name in waste_type_names:
            if waste_name in monthly_chart_data:
                weights = monthly_chart_data[waste_name]['weights']
                if sum(weights) > 0:  # Only consider types that have data
                    avg_weight = sum(weights) / len(weights)
                    variance = sum((w - avg_weight) ** 2 for w in weights) / len(weights)
                    variability_scores[waste_name] = variance
        
        if variability_scores:
            monthly_insights['most_consistent_type'] = min(variability_scores, key=variability_scores.get)
            monthly_insights['most_variable_type'] = max(variability_scores, key=variability_scores.get)
    
    # Prepare final monthly chart data for frontend
    monthly_tracking_data = {
        'months': formatted_months,
        'datasets': []
    }
    
    for waste_name in waste_type_names:
        if waste_name in monthly_chart_data and sum(monthly_chart_data[waste_name]['weights']) > 0:
            monthly_tracking_data['datasets'].append({
                'label': waste_name,
                'data': monthly_chart_data[waste_name]['weights'],
                'counts': monthly_chart_data[waste_name]['counts']
            })
    # ==================== END MONTHLY TRACKING ====================
    
    context = {
        'total_transactions': total_transactions,
        'total_weight': round(total_weight, 2),
        'total_points': round(total_points, 2),
        'avg_weight': round(avg_weight, 2),
        'unique_users': unique_users,
        'unique_barangays': unique_barangays,
        'co2_saved': round(co2_saved, 2),
        'trees_equivalent': round(trees_equivalent, 1),
        'most_popular_waste': most_popular_waste,
        'most_active_barangay': most_active_barangay,
        'waste_type_stats': waste_type_stats,
        'barangay_stats': barangay_stats,
        'top_collectors': top_collectors,
        'recent_transactions': recent_transactions,
        'all_barangays': all_barangays,
        'all_waste_types': all_waste_types,
        'waste_type_chart_data': json.dumps(waste_type_chart_data),
        'barangay_chart_data': json.dumps(barangay_chart_data),
        'waste_type_bar_data': json.dumps(waste_type_bar_data),
        'monthly_tracking_data': json.dumps(monthly_tracking_data),
        'monthly_insights': monthly_insights,
        'year_comparison': year_comparison,  # NEW: Year-over-year comparison data
        'filters': {
            'start_date': start_date,
            'end_date': end_date,
            'barangay_id': barangay_id,
            'waste_type_id': waste_type_id,
            'year_filter': year_filter,  # NEW: Year filter selection
        },
        'admin_user': request.admin_user,
        'timestamp': int(timezone.now().timestamp()),
        # Permission context for sidebar
        'can_manage_users': request.admin_user.can_manage_users,
        'can_manage_controls': request.admin_user.can_manage_controls,
        'can_manage_points': request.admin_user.can_manage_points,
        'can_manage_rewards': request.admin_user.can_manage_rewards,
        'can_manage_schedules': request.admin_user.can_manage_schedules,
        'can_manage_security': request.admin_user.can_manage_security,
        'can_manage_learning': request.admin_user.can_manage_learning,
        'can_manage_games': request.admin_user.can_manage_games,
    }
    
    return render(request, 'adminanalytics.html', context)


@admin_required
def download_excel_template(request):
    """
    Generate and download an Excel template for bulk waste data upload
    """
    if not EXCEL_AVAILABLE:
        messages.error(request, 'Excel support is not available. Please install openpyxl.')
        return redirect('cenro:waste_analytics')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Waste Collection Data"
    
    # Define headers
    headers = [
        'Transaction Date (YYYY-MM-DD)',
        'User ID or Family Code',
        'Waste Type',
        'Weight (kg)',
        'Barangay',
        'Notes (Optional)'
    ]
    
    # Style for headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = 25
    
    # Add example data
    example_data = [
        [
            '2025-01-15',
            'FAM001',
            'Plastic',
            '5.5',
            'Barangay 1',
            'Morning collection'
        ],
        [
            '2025-01-15',
            'user-uuid-here',
            'Paper',
            '3.2',
            'Barangay 2',
            ''
        ]
    ]
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num == 4:  # Weight column
                cell.number_format = '0.00'
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        ["Waste Collection Data Upload Template - Instructions"],
        [""],
        ["Column Descriptions:"],
        ["1. Transaction Date:", "Date of waste collection in YYYY-MM-DD format (e.g., 2025-01-15)"],
        ["2. User ID or Family Code:", "Either the user's UUID or their family code (e.g., FAM001)"],
        ["3. Waste Type:", "Type of waste collected (must match existing waste types)"],
        ["4. Weight (kg):", "Weight of waste in kilograms (decimal numbers allowed)"],
        ["5. Barangay:", "Name of the barangay (must match existing barangays)"],
        ["6. Notes:", "Optional notes about the transaction"],
        [""],
        ["Important Notes:"],
        ["• Do not modify the header row"],
        ["• Dates must be in YYYY-MM-DD format"],
        ["• Waste types and barangays must already exist in the system"],
        ["• Weight must be a positive number"],
        ["• Delete the example rows before uploading your data"],
        ["• Save the file as .xlsx format"],
    ]
    
    for row_num, instruction in enumerate(instructions, 1):
        for col_num, text in enumerate(instruction, 1):
            cell = ws_instructions.cell(row=row_num, column=col_num)
            cell.value = text
            if row_num == 1:
                cell.font = Font(bold=True, size=14, color='1F4E78')
            elif col_num == 1 and text.endswith(':'):
                cell.font = Font(bold=True, color='1F4E78')
            ws_instructions.column_dimensions[get_column_letter(col_num)].width = 40
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create response
    filename = f'waste_collection_template_{timezone.now().strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Log action
    log_admin_action(
        request.admin_user,
        None,
        'download_template',
        'Downloaded waste collection Excel template',
        request
    )
    
    return response


@admin_required
def upload_waste_data_excel(request):
    """
    Upload and process Excel file with waste collection data
    """
    if not EXCEL_AVAILABLE:
        return JsonResponse({
            'success': False,
            'error': 'Excel support is not available. Please install openpyxl.'
        }, status=400)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)
    
    if 'excel_file' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'No file uploaded'}, status=400)
    
    excel_file = request.FILES['excel_file']
    
    # Validate file extension
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({
            'success': False,
            'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'
        }, status=400)
    
    try:
        # Load workbook
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active
        
        # Skip header row
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        if not rows:
            return JsonResponse({
                'success': False,
                'error': 'No data found in Excel file'
            }, status=400)
        
        # Process rows
        success_count = 0
        error_count = 0
        errors = []
        
        for row_num, row in enumerate(rows, start=2):
            try:
                # Extract data
                transaction_date = row[0]
                user_identifier = row[1]
                waste_type_name = row[2]
                weight_kg = row[3]
                barangay_name = row[4]
                notes = row[5] if len(row) > 5 else ''
                
                # Validate required fields
                if not all([transaction_date, user_identifier, waste_type_name, weight_kg]):
                    errors.append(f'Row {row_num}: Missing required fields')
                    error_count += 1
                    continue
                
                # Parse date
                if isinstance(transaction_date, str):
                    transaction_date = datetime.strptime(transaction_date, '%Y-%m-%d').date()
                elif hasattr(transaction_date, 'date'):
                    transaction_date = transaction_date.date()
                
                # Find user
                user = None
                if str(user_identifier).startswith('FAM'):
                    # Find by family code
                    try:
                        family = Family.objects.get(family_code=user_identifier)
                        user = family.users.filter(is_active=True, status='approved').first()
                    except Family.DoesNotExist:
                        errors.append(f'Row {row_num}: Family code {user_identifier} not found')
                        error_count += 1
                        continue
                else:
                    # Find by user ID
                    try:
                        user = Users.objects.get(id=user_identifier, is_active=True, status='approved')
                    except Users.DoesNotExist:
                        errors.append(f'Row {row_num}: User ID {user_identifier} not found')
                        error_count += 1
                        continue
                
                if not user:
                    errors.append(f'Row {row_num}: No active user found for {user_identifier}')
                    error_count += 1
                    continue
                
                # Find waste type
                try:
                    waste_type = WasteType.objects.get(name__iexact=waste_type_name.strip())
                except WasteType.DoesNotExist:
                    errors.append(f'Row {row_num}: Waste type "{waste_type_name}" not found')
                    error_count += 1
                    continue
                
                # Find barangay
                barangay = None
                if barangay_name:
                    try:
                        barangay = Barangay.objects.get(name__iexact=barangay_name.strip())
                    except Barangay.DoesNotExist:
                        # Try to use user's barangay
                        if user.family and user.family.barangay:
                            barangay = user.family.barangay
                
                # Validate weight
                try:
                    weight_kg = float(weight_kg)
                    if weight_kg <= 0:
                        raise ValueError("Weight must be positive")
                except (ValueError, TypeError):
                    errors.append(f'Row {row_num}: Invalid weight value')
                    error_count += 1
                    continue
                
                # Calculate points
                total_points = weight_kg * waste_type.points_per_kg
                
                # Create transaction
                transaction = WasteTransaction.objects.create(
                    user=user,
                    waste_type=waste_type,
                    waste_kg=weight_kg,
                    total_points=total_points,
                    processed_by=request.admin_user,
                    barangay=barangay,
                    notes=notes,
                    created_at=timezone.make_aware(datetime.combine(transaction_date, datetime.min.time()))
                )
                
                # Manually set transaction_date
                transaction.transaction_date = transaction_date
                transaction.save(update_fields=['transaction_date'])
                
                # Update user points
                user.total_points += total_points
                user.save(update_fields=['total_points'])
                
                # Update family points
                if user.family:
                    user.family.total_family_points += total_points
                    user.family.save(update_fields=['total_family_points'])
                
                success_count += 1
                
            except Exception as e:
                errors.append(f'Row {row_num}: {str(e)}')
                error_count += 1
                logger.error(f"Error processing row {row_num}: {str(e)}")
        
        # Log action
        log_admin_action(
            request.admin_user,
            None,
            'excel_upload',
            f'Uploaded waste data via Excel: {success_count} success, {error_count} errors',
            request
        )
        
        response_data = {
            'success': True,
            'success_count': success_count,
            'error_count': error_count,
            'total_rows': len(rows),
            'message': f'Processed {success_count} transactions successfully'
        }
        
        if errors:
            response_data['errors'] = errors[:50]  # Limit to first 50 errors
        
        return JsonResponse(response_data)
        
    except Exception as e:
        logger.error(f"Error uploading Excel file: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Failed to process Excel file: {str(e)}'
        }, status=500)


@admin_required
def export_waste_data_excel(request):
    """
    Export waste collection data to Excel with current filters
    """
    if not EXCEL_AVAILABLE:
        messages.error(request, 'Excel support is not available. Please install openpyxl.')
        return redirect('cenro:waste_analytics')
    
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    barangay_id = request.GET.get('barangay')
    waste_type_id = request.GET.get('waste_type')
    
    # Base queryset
    transactions = WasteTransaction.objects.select_related(
        'user', 'waste_type', 'barangay', 'processed_by', 'user__family'
    ).all()
    
    # Apply filters
    if start_date:
        transactions = transactions.filter(transaction_date__gte=start_date)
    if end_date:
        transactions = transactions.filter(transaction_date__lte=end_date)
    if barangay_id:
        transactions = transactions.filter(barangay_id=barangay_id)
    if waste_type_id:
        transactions = transactions.filter(waste_type_id=waste_type_id)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Waste Collection Data"
    
    # Define headers
    headers = [
        'Date', 'User Name', 'Family Code', 'Waste Type', 
        'Weight (kg)', 'Points Earned', 'Barangay', 
        'Processed By', 'Notes', 'Transaction Time'
    ]
    
    # Style for headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment
        cell.border = border
    
    # Write data
    for row_num, transaction in enumerate(transactions, 2):
        data = [
            transaction.transaction_date.strftime('%Y-%m-%d'),
            transaction.user.full_name,
            transaction.user.family.family_code if transaction.user.family else 'N/A',
            transaction.waste_type.name,
            transaction.waste_kg,
            transaction.total_points,
            transaction.barangay.name if transaction.barangay else 'N/A',
            transaction.processed_by.full_name if transaction.processed_by else 'System',
            transaction.notes or '',
            transaction.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            # Format numbers
            if col_num in [5, 6]:  # Weight and Points columns
                cell.number_format = '0.00'
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 18
    
    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    
    total_weight = transactions.aggregate(Sum('waste_kg'))['waste_kg__sum'] or 0
    total_points = transactions.aggregate(Sum('total_points'))['total_points__sum'] or 0
    total_transactions = transactions.count()
    
    summary_data = [
        ["Waste Collection Summary Report"],
        [""],
        ["Report Generated:", timezone.now().strftime('%Y-%m-%d %H:%M:%S')],
        ["Total Transactions:", total_transactions],
        ["Total Weight (kg):", round(total_weight, 2)],
        ["Total Points:", round(total_points, 2)],
        [""],
        ["Filters Applied:"],
        ["Start Date:", start_date or 'All'],
        ["End Date:", end_date or 'All'],
        ["Barangay:", request.GET.get('barangay_name', 'All')],
        ["Waste Type:", request.GET.get('waste_type_name', 'All')],
    ]
    
    for row_num, row_data in enumerate(summary_data, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_num, column=col_num)
            cell.value = value
            if row_num == 1:
                cell.font = Font(bold=True, size=14, color='1F4E78')
            elif col_num == 1 and row_num > 2:
                cell.font = Font(bold=True)
            ws_summary.column_dimensions[get_column_letter(col_num)].width = 30
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create response
    filename = f'waste_collection_data_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Log action
    log_admin_action(
        request.admin_user,
        None,
        'export_excel',
        f'Exported {total_transactions} waste transactions to Excel',
        request
    )
    
    return response


def _generate_pie_chart(waste_type_stats, total_weight):
    """Generate a pie chart for waste type distribution"""
    if not MATPLOTLIB_AVAILABLE:
        logger.warning("Matplotlib not available for chart generation")
        return None
    
    try:
        # Convert QuerySet to list and validate
        stats_list = list(waste_type_stats[:8]) if waste_type_stats else []
        
        if not stats_list:
            logger.info("No waste type data available for pie chart")
            return None
        
        # Prepare data
        labels = [item['waste_type__name'] or 'Unknown' for item in stats_list]
        sizes = [float(item['total_weight']) for item in stats_list]
        
        # Validate data
        if not labels or not sizes or sum(sizes) == 0:
            logger.warning("Invalid data for pie chart generation")
            return None
        
        # Create color palette
        colors_palette = ['#667eea', '#764ba2', '#f093fb', '#4facfe', 
                         '#43e97b', '#38f9d7', '#fa709a', '#fee140']
        
        # Create figure with better aspect ratio for PDF (wider)
        fig, ax = plt.subplots(figsize=(8, 5), facecolor='white')
        
        # Create pie chart with better text size
        wedges, texts, autotexts = ax.pie(
            sizes, 
            labels=labels, 
            autopct='%1.1f%%',
            colors=colors_palette[:len(sizes)],
            startangle=90,
            textprops={'fontsize': 11, 'weight': 'bold'}
        )
        
        # Enhance text for better readability
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_fontsize(10)
            autotext.set_weight('bold')
        
        ax.set_title('Waste Type Distribution', fontsize=14, weight='bold', pad=15)
        
        # Save to BytesIO
        img_buffer = io.BytesIO()
        plt.tight_layout()
        plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight', facecolor='white')
        plt.close(fig)  # Close specific figure
        img_buffer.seek(0)
        
        logger.info(f"Pie chart generated successfully with {len(labels)} items")
        return img_buffer
    except Exception as e:
        logger.error(f"Error generating pie chart: {e}")
        return None


def _generate_bar_chart(barangay_stats):
    """Generate a horizontal bar chart for barangay collection"""
    if not MATPLOTLIB_AVAILABLE:
        logger.warning("Matplotlib not available for bar chart generation")
        return None
    
    try:
        # Convert QuerySet to list and validate
        stats_list = list(barangay_stats[:10]) if barangay_stats else []
        
        if not stats_list:
            logger.info("No barangay data available for bar chart")
            return None
        
        # Prepare data (top 10 barangays)
        labels = [item['barangay__name'] or 'Unknown' for item in stats_list]
        values = [float(item['total_weight']) for item in stats_list]
        
        # Validate data
        if not labels or not values:
            logger.warning("Invalid data for bar chart generation")
            return None
        
        # Reverse for better display (highest at top)
        labels = labels[::-1]
        values = values[::-1]
        
        # Create figure with better dimensions for horizontal bars
        fig, ax = plt.subplots(figsize=(8, 5), facecolor='white')
        
        # Create horizontal bar chart
        bars = ax.barh(labels, values, color='#667eea', edgecolor='#764ba2', linewidth=1.5, height=0.6)
        
        # Add value labels on bars with better formatting
        for i, (bar, value) in enumerate(zip(bars, values)):
            width = bar.get_width()
            ax.text(width, bar.get_y() + bar.get_height()/2, 
                   f' {value:.1f} kg',
                   ha='left', va='center', fontsize=10, weight='bold')
        
        ax.set_xlabel('Weight (kg)', fontsize=11, weight='bold')
        ax.set_title('Collection by Barangay', fontsize=14, weight='bold', pad=15)
        ax.grid(axis='x', alpha=0.3, linestyle='--')
        ax.tick_params(axis='y', labelsize=10)
        ax.tick_params(axis='x', labelsize=9)
        
        # Save to BytesIO
        img_buffer = io.BytesIO()
        plt.tight_layout()
        plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight', facecolor='white')
        plt.close(fig)  # Close specific figure
        img_buffer.seek(0)
        
        logger.info(f"Bar chart generated successfully with {len(labels)} barangays")
        return img_buffer
    except Exception as e:
        logger.error(f"Error generating bar chart: {e}")
        return None


def _generate_comparison_chart(waste_type_stats):
    """Generate a grouped bar chart comparing weight and transactions"""
    if not MATPLOTLIB_AVAILABLE:
        logger.warning("Matplotlib not available for comparison chart generation")
        return None
    
    try:
        # Convert QuerySet to list and validate
        stats_list = list(waste_type_stats[:8]) if waste_type_stats else []
        
        if not stats_list:
            logger.info("No waste type data available for comparison chart")
            return None
        
        # Prepare data (top 8 waste types)
        labels = [item['waste_type__name'] or 'Unknown' for item in stats_list]
        weights = [float(item['total_weight']) for item in stats_list]
        counts = [item['count'] for item in stats_list]
        
        # Validate data
        if not labels or not weights or not counts:
            logger.warning("Invalid data for comparison chart generation")
            return None
        
        # Create figure with better dimensions for comparison chart
        fig, ax1 = plt.subplots(figsize=(9, 5), facecolor='white')
        
        x = range(len(labels))
        width = 0.35
        
        # Plot weight bars with better styling
        bars1 = ax1.bar([i - width/2 for i in x], weights, width, 
                        label='Weight (kg)', color='#667eea', edgecolor='#764ba2', linewidth=1.5)
        ax1.set_xlabel('Waste Type', fontsize=11, weight='bold')
        ax1.set_ylabel('Weight (kg)', fontsize=11, weight='bold', color='#667eea')
        ax1.tick_params(axis='y', labelcolor='#667eea', labelsize=10)
        ax1.set_xticks(x)
        ax1.set_xticklabels(labels, rotation=45, ha='right', fontsize=10)
        
        # Create second y-axis for transactions
        ax2 = ax1.twinx()
        bars2 = ax2.bar([i + width/2 for i in x], counts, width,
                        label='Transactions', color='#10b981', edgecolor='#059669', linewidth=1.5)
        ax2.set_ylabel('Number of Transactions', fontsize=11, weight='bold', color='#10b981')
        ax2.tick_params(axis='y', labelcolor='#10b981', labelsize=10)
        
        # Title and legend
        ax1.set_title('Waste Type Comparison: Weight vs Transactions', fontsize=14, weight='bold', pad=15)
        
        # Add legends with better positioning
        ax1.legend(loc='upper left', fontsize=10)
        ax2.legend(loc='upper right', fontsize=10)
        
        ax1.grid(axis='y', alpha=0.3, linestyle='--')
        
        # Save to BytesIO
        img_buffer = io.BytesIO()
        plt.tight_layout()
        plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight', facecolor='white')
        plt.close(fig)  # Close specific figure
        img_buffer.seek(0)
        
        logger.info(f"Comparison chart generated successfully with {len(labels)} items")
        return img_buffer
    except Exception as e:
        logger.error(f"Error generating comparison chart: {e}")
        return None


def _generate_monthly_tracking_chart(transactions):
    """Generate a line chart for monthly waste collection trends by type"""
    if not MATPLOTLIB_AVAILABLE:
        logger.warning("Matplotlib not available for monthly tracking chart generation")
        return None
    
    try:
        # Aggregate transactions by month and waste type
        monthly_waste_data = transactions.annotate(
            month=TruncMonth('transaction_date')
        ).values('month', 'waste_type__name').annotate(
            total_weight=Sum('waste_kg'),
            transaction_count=Count('id')
        ).order_by('month', 'waste_type__name')
        
        if not monthly_waste_data:
            logger.info("No monthly data available for tracking chart")
            return None
        
        # Organize data
        monthly_chart_data = {}
        all_months = []
        waste_types = set()
        
        for item in monthly_waste_data:
            month_str = item['month'].strftime('%b %Y') if item['month'] else 'Unknown'
            waste_name = item['waste_type__name']
            waste_types.add(waste_name)
            
            if month_str not in all_months:
                all_months.append(month_str)
            
            if waste_name not in monthly_chart_data:
                monthly_chart_data[waste_name] = {}
            
            monthly_chart_data[waste_name][month_str] = float(item['total_weight'])
        
        if not all_months or not waste_types:
            logger.warning("Invalid data for monthly tracking chart generation")
            return None
        
        # Create figure with optimal size for PDF
        fig, ax = plt.subplots(figsize=(11, 6), facecolor='white')
        
        # Professional color palette
        colors = ['#667eea', '#10b981', '#f59e0b', '#ef4444', '#3b82f6', 
                  '#ec4899', '#8b5cf6', '#14b8a6', '#f97316', '#06b6d4']
        
        # Plot each waste type as a line
        for idx, waste_type in enumerate(sorted(waste_types)):
            weights = [monthly_chart_data[waste_type].get(month, 0) for month in all_months]
            color = colors[idx % len(colors)]
            
            ax.plot(all_months, weights, marker='o', linewidth=3, markersize=7,
                   label=waste_type, color=color, markerfacecolor=color,
                   markeredgecolor='white', markeredgewidth=2)
        
        # Styling
        ax.set_xlabel('Collection Period (Month)', fontsize=12, weight='bold', color='#374151')
        ax.set_ylabel('Weight Collected (kg)', fontsize=12, weight='bold', color='#374151')
        ax.set_title('Monthly Waste Collection Trends by Type', fontsize=15, weight='bold', pad=20, color='#1f2937')
        
        # Grid and legend
        ax.grid(axis='y', alpha=0.3, linestyle='--', linewidth=0.8)
        ax.legend(loc='upper left', fontsize=9, framealpha=0.95, edgecolor='#e5e7eb')
        
        # Rotate x-axis labels for better readability
        plt.xticks(rotation=45, ha='right', fontsize=10)
        plt.yticks(fontsize=10)
        
        # Add subtle background
        ax.set_facecolor('#fafafa')
        
        # Save to BytesIO
        img_buffer = io.BytesIO()
        plt.tight_layout()
        plt.savefig(img_buffer, format='png', dpi=200, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        img_buffer.seek(0)
        
        logger.info(f"Monthly tracking chart generated with {len(all_months)} months and {len(waste_types)} waste types")
        return img_buffer
    except Exception as e:
        logger.error(f"Error generating monthly tracking chart: {e}")
        return None


def _calculate_monthly_insights(transactions):
    """Calculate monthly insights for PDF report"""
    try:
        # Aggregate by month
        monthly_waste_data = transactions.annotate(
            month=TruncMonth('transaction_date')
        ).values('month', 'waste_type__name').annotate(
            total_weight=Sum('waste_kg'),
            transaction_count=Count('id')
        ).order_by('month', 'waste_type__name')
        
        if not monthly_waste_data:
            return None
        
        # Calculate insights
        monthly_totals = {}
        waste_type_data = {}
        all_months = []
        
        for item in monthly_waste_data:
            if item['month']:
                month_str = item['month'].strftime('%b %Y')
                waste_name = item['waste_type__name']
                weight = float(item['total_weight'])
                
                if month_str not in all_months:
                    all_months.append(month_str)
                
                monthly_totals[month_str] = monthly_totals.get(month_str, 0) + weight
                
                if waste_name not in waste_type_data:
                    waste_type_data[waste_name] = []
                waste_type_data[waste_name].append(weight)
        
        if not all_months:
            return None
        
        insights = {
            'total_months': len(all_months),
            'peak_month': None,
            'peak_month_weight': 0,
            'trend': 'stable',
            'growth_rate': 0,
            'most_consistent_type': None,
            'most_variable_type': None,
        }
        
        # Find peak month
        for month, total in monthly_totals.items():
            if total > insights['peak_month_weight']:
                insights['peak_month_weight'] = total
                insights['peak_month'] = month
        
        # Calculate trend
        if len(all_months) >= 2:
            mid_point = len(all_months) // 2
            month_values = list(monthly_totals.values())
            first_half_avg = sum(month_values[:mid_point]) / mid_point if mid_point > 0 else 0
            second_half_avg = sum(month_values[mid_point:]) / (len(month_values) - mid_point) if (len(month_values) - mid_point) > 0 else 0
            
            if second_half_avg > first_half_avg * 1.1:
                insights['trend'] = 'increasing'
                insights['growth_rate'] = ((second_half_avg - first_half_avg) / first_half_avg * 100) if first_half_avg > 0 else 0
            elif second_half_avg < first_half_avg * 0.9:
                insights['trend'] = 'decreasing'
                insights['growth_rate'] = ((first_half_avg - second_half_avg) / first_half_avg * 100) if first_half_avg > 0 else 0
        
        # Find most consistent and variable
        # Only calculate if we have at least 2 months of data (need multiple data points for meaningful variance)
        variability_scores = {}
        for waste_name, weights in waste_type_data.items():
            # Need at least 2 data points to calculate meaningful variance
            if len(weights) >= 2 and sum(weights) > 0:
                avg = sum(weights) / len(weights)
                variance = sum((w - avg) ** 2 for w in weights) / len(weights)
                variability_scores[waste_name] = variance
        
        if variability_scores:
            # Only assign if there are at least 2 different waste types with meaningful variance
            if len(variability_scores) >= 2:
                # Check if variances are actually different (not all zero)
                unique_variances = set(variability_scores.values())
                if len(unique_variances) > 1:
                    insights['most_consistent_type'] = min(variability_scores, key=variability_scores.get)
                    insights['most_variable_type'] = max(variability_scores, key=variability_scores.get)
                else:
                    # All variances are the same - can't determine most/least variable
                    insights['most_consistent_type'] = None
                    insights['most_variable_type'] = None
            elif len(variability_scores) == 1:
                # If only one waste type has enough data points
                single_type = list(variability_scores.keys())[0]
                insights['most_consistent_type'] = single_type
                insights['most_variable_type'] = None
        
        return insights
    except Exception as e:
        logger.error(f"Error calculating monthly insights: {e}")
        return None


@admin_required
def export_waste_data_pdf(request):
    """
    Export waste collection data to PDF with charts and tables
    """
    if not PDF_AVAILABLE:
        messages.error(request, 'PDF support is not available. Please install reportlab.')
        return redirect('cenro:waste_analytics')
    
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    barangay_id = request.GET.get('barangay')
    waste_type_id = request.GET.get('waste_type')
    
    # Base queryset
    transactions = WasteTransaction.objects.select_related(
        'user', 'waste_type', 'barangay', 'processed_by', 'user__family'
    ).all()
    
    # Apply filters
    if start_date:
        transactions = transactions.filter(transaction_date__gte=start_date)
    if end_date:
        transactions = transactions.filter(transaction_date__lte=end_date)
    if barangay_id:
        transactions = transactions.filter(barangay_id=barangay_id)
    if waste_type_id:
        transactions = transactions.filter(waste_type_id=waste_type_id)
    
    # Calculate statistics
    total_weight = transactions.aggregate(Sum('waste_kg'))['waste_kg__sum'] or 0
    total_points = transactions.aggregate(Sum('total_points'))['total_points__sum'] or 0
    total_transactions = transactions.count()
    avg_weight = transactions.aggregate(Avg('waste_kg'))['waste_kg__avg'] or 0
    
    # Waste type breakdown
    waste_type_stats = transactions.values(
        'waste_type__name'
    ).annotate(
        total_weight=Sum('waste_kg'),
        count=Count('id')
    ).order_by('-total_weight')
    
    # Barangay breakdown
    barangay_stats = transactions.values(
        'barangay__name'
    ).annotate(
        total_weight=Sum('waste_kg'),
        count=Count('id')
    ).order_by('-total_weight')
    
    # Create PDF with better margins for landscape
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=40,
        leftMargin=40,
        topMargin=35,
        bottomMargin=35
    )
    
    # Container for PDF elements
    elements = []
    styles = getSampleStyleSheet()
    
    # Enhanced title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=26,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=20,
        alignment=1,  # Center
        fontName='Helvetica-Bold'
    )
    
    # Enhanced subtitle/heading style
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2C5282'),
        spaceAfter=8,
        spaceBefore=5,
        fontName='Helvetica-Bold'
    )
    elements.append(Paragraph("Waste Collection Analytics Report", title_style))
    elements.append(Spacer(1, 8))
    
    # Report Info with better formatting
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#4A5568'),
        alignment=1,  # Center
        spaceAfter=15
    )
    info_text = f"""
    <b>Report Generated:</b> {timezone.now().strftime('%B %d, %Y at %I:%M %p')}<br/>
    <b>Date Range:</b> {start_date or 'All'} to {end_date or 'All'}<br/>
    <b>Generated By:</b> {request.admin_user.full_name}
    """
    elements.append(Paragraph(info_text, info_style))
    elements.append(Spacer(1, 15))
    
    # Summary Statistics Table
    summary_data = [
        ['Metric', 'Value'],
        ['Total Transactions', f'{total_transactions:,}'],
        ['Total Weight Collected', f'{total_weight:,.2f} kg'],
        ['Total Points Awarded', f'{total_points:,.2f}'],
        ['Average Weight per Transaction', f'{avg_weight:.2f} kg'],
        ['CO₂ Saved (estimated)', f'{(total_weight * 0.5):,.1f} kg'],
        ['Trees Equivalent', f'{(total_weight * 0.02):,.1f}'],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey])
    ]))
    
    elements.append(Paragraph("<b>Summary Statistics</b>", heading_style))
    elements.append(Spacer(1, 8))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Generate and add charts
    logger.info(f"MATPLOTLIB_AVAILABLE: {MATPLOTLIB_AVAILABLE}")
    
    if MATPLOTLIB_AVAILABLE:
        logger.info(f"Attempting to generate charts. waste_type_stats count: {len(list(waste_type_stats)) if waste_type_stats else 0}")
        logger.info(f"barangay_stats count: {len(list(barangay_stats)) if barangay_stats else 0}")
        
        # Add page break before charts section for better layout
        elements.append(PageBreak())
        
        # Waste Type Pie Chart
        try:
            pie_chart_buffer = _generate_pie_chart(waste_type_stats, total_weight)
            if pie_chart_buffer:
                logger.info("Pie chart generated, adding to PDF")
                elements.append(Paragraph("Waste Type Distribution Chart", heading_style))
                elements.append(Spacer(1, 5))
                
                # Add DATA-SPECIFIC explanation with actual results
                explanation_style = ParagraphStyle(
                    'ExplanationStyle',
                    parent=styles['Normal'],
                    fontSize=9,
                    textColor=colors.HexColor('#4A5568'),
                    leading=12,
                    spaceAfter=10,
                    leftIndent=10
                )
                
                # Build dynamic explanation from actual data
                stats_list = list(waste_type_stats[:3])
                if stats_list:
                    first = stats_list[0]
                    first_percentage = (float(first['total_weight']) / total_weight * 100) if total_weight > 0 else 0
                    explanation_parts = [
                        f"<b>Understanding This Chart:</b> The results show that <b>{first['waste_type__name']}</b> has the highest collection volume with "
                        f"<b>{first['total_weight']:.2f} kg</b> ({first_percentage:.0f}% of total)"
                    ]
                    
                    if len(stats_list) > 1:
                        second = stats_list[1]
                        explanation_parts.append(f", followed by <b>{second['waste_type__name']}</b> at <b>{second['total_weight']:.2f} kg</b>")
                    
                    if len(stats_list) > 2:
                        third = stats_list[2]
                        explanation_parts.append(f" and <b>{third['waste_type__name']}</b> at <b>{third['total_weight']:.2f} kg</b>")
                    
                    explanation_parts.append(". This distribution reveals which waste types dominate your collection program.")
                    explanation_text = ''.join(explanation_parts)
                else:
                    explanation_text = "<b>Understanding This Chart:</b> Proportional distribution of waste types as percentages of total collection."
                
                elements.append(Paragraph(explanation_text, explanation_style))
                elements.append(Spacer(1, 5))
                
                # Larger size with proper aspect ratio (8:5 ratio maintained)
                img = Image(pie_chart_buffer, width=6.4*inch, height=4*inch)
                img.hAlign = 'CENTER'
                elements.append(img)
                elements.append(Spacer(1, 15))
            else:
                logger.warning("Pie chart buffer is None")
        except Exception as e:
            logger.error(f"Error adding pie chart to PDF: {e}", exc_info=True)
        
        # Barangay Bar Chart
        try:
            bar_chart_buffer = _generate_bar_chart(barangay_stats)
            if bar_chart_buffer:
                logger.info("Bar chart generated, adding to PDF")
                elements.append(Paragraph("Collection by Barangay Chart", heading_style))
                elements.append(Spacer(1, 5))
                
                # Add DATA-SPECIFIC explanation with actual results
                explanation_style = ParagraphStyle(
                    'ExplanationStyle',
                    parent=styles['Normal'],
                    fontSize=9,
                    textColor=colors.HexColor('#4A5568'),
                    leading=12,
                    spaceAfter=10,
                    leftIndent=10
                )
                
                # Build dynamic explanation from actual data
                barangay_list = list(barangay_stats[:3])
                if barangay_list:
                    first = barangay_list[0]
                    explanation_parts = [
                        f"<b>Understanding This Chart:</b> The results show that <b>{first['barangay__name'] or 'Unknown Barangay'}</b> has the highest "
                        f"collection performance with <b>{first['total_weight']:.2f} kg</b> collected across <b>{first['count']}</b> transactions"
                    ]
                    
                    if len(barangay_list) > 1:
                        second = barangay_list[1]
                        explanation_parts.append(f". This is followed by <b>{second['barangay__name'] or 'Unknown'}</b> with <b>{second['total_weight']:.2f} kg</b>")
                    
                    if len(barangay_list) > 2:
                        third = barangay_list[2]
                        explanation_parts.append(f" and <b>{third['barangay__name'] or 'Unknown'}</b> with <b>{third['total_weight']:.2f} kg</b>")
                    
                    explanation_parts.append(".")
                    explanation_text = ''.join(explanation_parts)
                else:
                    explanation_text = "<b>Understanding This Chart:</b> Comparative collection performance across barangays ranked by total weight."
                
                elements.append(Paragraph(explanation_text, explanation_style))
                elements.append(Spacer(1, 5))
                
                # Larger size with proper aspect ratio (8:5 ratio maintained)
                img = Image(bar_chart_buffer, width=6.4*inch, height=4*inch)
                img.hAlign = 'CENTER'
                elements.append(img)
                elements.append(Spacer(1, 15))
            else:
                logger.warning("Bar chart buffer is None")
        except Exception as e:
            logger.error(f"Error adding bar chart to PDF: {e}", exc_info=True)
        
        # Add page break before comparison chart for cleaner layout
        elements.append(PageBreak())
        
        # Waste Type Comparison Chart
        try:
            comparison_chart_buffer = _generate_comparison_chart(waste_type_stats)
            if comparison_chart_buffer:
                logger.info("Comparison chart generated, adding to PDF")
                elements.append(Paragraph("Waste Type Comparison Chart", heading_style))
                elements.append(Spacer(1, 5))
                
                # Add DATA-SPECIFIC explanation with actual results
                explanation_style = ParagraphStyle(
                    'ExplanationStyle',
                    parent=styles['Normal'],
                    fontSize=9,
                    textColor=colors.HexColor('#4A5568'),
                    leading=12,
                    spaceAfter=10,
                    leftIndent=10
                )
                
                # Build dynamic explanation from actual data
                comparison_list = list(waste_type_stats[:2])
                if comparison_list:
                    first = comparison_list[0]
                    explanation_parts = [
                        f"<b>Understanding This Chart:</b> The comparison reveals that <b>{first['waste_type__name']}</b> leads in both total weight "
                        f"(<b>{first['total_weight']:.2f} kg</b>) and transactions (<b>{first['count']}</b>)"
                    ]
                    
                    if len(comparison_list) > 1:
                        second = comparison_list[1]
                        explanation_parts.append(
                            f". <b>{second['waste_type__name']}</b> follows with <b>{second['total_weight']:.2f} kg</b> "
                            f"across <b>{second['count']}</b> transactions"
                        )
                    
                    explanation_parts.append(". This dual view shows which waste types contribute the most volume and which generate the most collection activity.")
                    explanation_text = ''.join(explanation_parts)
                else:
                    explanation_text = "<b>Understanding This Chart:</b> Dual comparison of total weight versus transaction count for each waste type."
                
                elements.append(Paragraph(explanation_text, explanation_style))
                elements.append(Spacer(1, 5))
                
                # Larger size with proper aspect ratio (9:5 ratio maintained)
                img = Image(comparison_chart_buffer, width=7.2*inch, height=4*inch)
                img.hAlign = 'CENTER'
                elements.append(img)
                elements.append(Spacer(1, 20))
            else:
                logger.warning("Comparison chart buffer is None")
        except Exception as e:
            logger.error(f"Error adding comparison chart to PDF: {e}", exc_info=True)
        
        # Monthly Tracking Chart (NEW ADDITION)
        try:
            monthly_chart_buffer = _generate_monthly_tracking_chart(transactions)
            if monthly_chart_buffer:
                logger.info("Monthly tracking chart generated, adding to PDF")
                elements.append(PageBreak())
                elements.append(Paragraph("Monthly Waste Collection Trends by Type", heading_style))
                elements.append(Spacer(1, 8))
                
                # Add DATA-SPECIFIC explanation
                explanation_style = ParagraphStyle(
                    'ExplanationStyle',
                    parent=styles['Normal'],
                    fontSize=9,
                    textColor=colors.HexColor('#4A5568'),
                    leading=12,
                    spaceAfter=10,
                    leftIndent=10
                )
                
                # Calculate monthly insights for explanation
                monthly_insights = _calculate_monthly_insights(transactions)
                if monthly_insights and monthly_insights.get('peak_month'):
                    explanation_text = (
                        f"<b>Understanding This Chart:</b> Monthly waste collection trends showing {monthly_insights['total_months']} months of data. "
                        f"Peak collection occurred in <b>{monthly_insights['peak_month']}</b> with <b>{monthly_insights['peak_month_weight']:.1f} kg</b> collected. "
                        f"Overall trend shows <b>{monthly_insights['trend']}</b> collection patterns"
                    )
                    if monthly_insights['growth_rate'] > 0:
                        explanation_text += f" with {monthly_insights['growth_rate']:.1f}% change"
                    explanation_text += ". Each line represents a different waste type's monthly collection volume over time."
                else:
                    explanation_text = (
                        "<b>Understanding This Chart:</b> Monthly waste collection trends for each waste type over time. "
                        "Each colored line represents a different waste type's collection volume."
                    )
                
                elements.append(Paragraph(explanation_text, explanation_style))
                elements.append(Spacer(1, 8))
                
                # Add chart image (larger for better readability)
                img = Image(monthly_chart_buffer, width=9*inch, height=5*inch)
                img.hAlign = 'CENTER'
                elements.append(img)
                elements.append(Spacer(1, 15))
                
                # Add monthly insights if available
                monthly_insights = _calculate_monthly_insights(transactions)
                if monthly_insights:
                    insights_style = ParagraphStyle(
                        'InsightsStyle',
                        parent=styles['Normal'],
                        fontSize=9,
                        textColor=colors.HexColor('#1F4E78'),
                        leading=12,
                        spaceAfter=8,
                        leftIndent=10
                    )
                    
                    elements.append(Paragraph("<b>Key Monthly Insights:</b>", heading_style))
                    elements.append(Spacer(1, 5))
                    
                    if monthly_insights['peak_month']:
                        insight_text = f"• <b>Peak Collection Month:</b> {monthly_insights['peak_month']} with {monthly_insights['peak_month_weight']:.1f} kg collected"
                        elements.append(Paragraph(insight_text, insights_style))
                    
                    trend_emoji = {'increasing': '📈', 'decreasing': '📉', 'stable': '➡️'}
                    trend_text = f"• <b>Overall Trend:</b> {trend_emoji.get(monthly_insights['trend'], '')} {monthly_insights['trend'].capitalize()}"
                    if monthly_insights['growth_rate'] > 0:
                        trend_text += f" ({monthly_insights['growth_rate']:.1f}% change)"
                    elements.append(Paragraph(trend_text, insights_style))
                    
                    if monthly_insights['most_consistent_type']:
                        elements.append(Paragraph(f"• <b>Most Consistent Collection:</b> {monthly_insights['most_consistent_type']} (stable month-to-month)", insights_style))
                    
                    if monthly_insights['most_variable_type']:
                        elements.append(Paragraph(f"• <b>Most Variable Collection:</b> {monthly_insights['most_variable_type']} (seasonal fluctuations)", insights_style))
                    
                    elements.append(Spacer(1, 10))
            else:
                logger.warning("Monthly tracking chart buffer is None")
        except Exception as e:
            logger.error(f"Error adding monthly tracking chart to PDF: {e}", exc_info=True)
    else:
        logger.warning("Matplotlib not available, skipping chart generation")
    
    # Add page break before data tables
    elements.append(PageBreak())
    
    # Waste Type Breakdown
    if waste_type_stats:
        waste_type_data = [['Waste Type', 'Total Weight (kg)', 'Transactions']]
        for item in waste_type_stats[:10]:
            waste_type_data.append([
                item['waste_type__name'] or 'Unknown',
                f"{item['total_weight']:.2f}",
                str(item['count'])
            ])
        
        waste_type_table = Table(waste_type_data, colWidths=[3*inch, 2*inch, 2*inch])
        waste_type_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#70AD47')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey])
        ]))
        
        elements.append(Paragraph("Waste Type Breakdown", heading_style))
        elements.append(Spacer(1, 8))
        elements.append(waste_type_table)
        elements.append(Spacer(1, 20))
    
    # Barangay Breakdown
    if barangay_stats:
        barangay_data = [['Barangay', 'Total Weight (kg)', 'Transactions']]
        for item in barangay_stats[:10]:
            barangay_data.append([
                item['barangay__name'] or 'Unknown',
                f"{item['total_weight']:.2f}",
                str(item['count'])
            ])
        
        barangay_table = Table(barangay_data, colWidths=[3*inch, 2*inch, 2*inch])
        barangay_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ED7D31')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey])
        ]))
        
        elements.append(Paragraph("Barangay Breakdown", heading_style))
        elements.append(Spacer(1, 8))
        elements.append(barangay_table)
        elements.append(Spacer(1, 20))
    
    # Recent Transactions (last 20)
    recent_transactions = transactions.order_by('-created_at')[:20]
    if recent_transactions:
        elements.append(PageBreak())
        elements.append(Paragraph("Recent Transactions", heading_style))
        elements.append(Spacer(1, 8))
        
        transaction_data = [['Date', 'User', 'Waste Type', 'Weight (kg)', 'Points']]
        for t in recent_transactions:
            transaction_data.append([
                t.transaction_date.strftime('%Y-%m-%d'),
                t.user.full_name[:20],
                t.waste_type.name[:15],
                f"{t.waste_kg:.2f}",
                f"{t.total_points:.0f}"
            ])
        
        transaction_table = Table(transaction_data, colWidths=[1.5*inch, 2.5*inch, 2*inch, 1.2*inch, 1*inch])
        transaction_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5B9BD5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey])
        ]))
        
        elements.append(transaction_table)
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF content
    pdf_content = buffer.getvalue()
    buffer.close()
    
    # Create response
    filename = f'waste_analytics_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response = HttpResponse(pdf_content, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Log action
    log_admin_action(
        request.admin_user,
        None,
        'export_pdf',
        f'Exported {total_transactions} waste transactions to PDF',
        request
    )
    
    return response
